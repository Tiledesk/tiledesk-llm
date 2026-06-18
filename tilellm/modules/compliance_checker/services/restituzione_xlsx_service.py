"""
RestituzioneXlsxService — renders compliance check results into the standardized
*restituzione* (output) workbook aligned to the market template
`ipotesi tabella criteri.xlsx`.

Columns (exact market labels, plus the "Operatore Economico" column requested for
the massive multi-operator evaluation):

    Operatore Economico
    Criterio
    Tipo criterio
    Modalità
    Punteggio previsto
    Presenza del requisito (per il tipo conformità)
    Punteggio assegnato (per i tipi tabellari e discrezionali)
    Livello di confidenza dell'analisi
    Motivazioni e argomentazioni a supporto del punteggio assegnato (per i tabellari e discrezionali)
    Nome del/dei documenti utilizzati ove è trattato il criterio
    n. pagina/e del/dei documento/i ove è trattato il criterio
    Note e/o descrizione aggiuntiva a supporto dell'analisi condotta

`build_workbook` accepts a LIST of reports so the same renderer serves both the
single-operator case (Fase 1) and the massive multi-operator case (Fase 2): one
sheet per lot, rows ordered **by criterion, then operator**.
"""
import io
import logging
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tilellm.modules.compliance_checker.models import _DEFAULT_JUDGMENT_MAP
from tilellm.modules.compliance_checker.models_v2 import ComplianceReportV2
from tilellm.modules.compliance_checker.services import xlsx_taxonomy as tax

logger = logging.getLogger(__name__)

# Exact market-template headers (+ Operatore Economico).
COL_OPERATOR = "Operatore Economico"
COL_CRITERIO = "Criterio"
COL_TIPO = "Tipo criterio"
COL_MODE = "Modalità"
COL_MAX_POINTS = "Punteggio previsto"
COL_PRESENZA = "Presenza del requisito (per il tipo conformità)"
COL_SCORE = "Punteggio assegnato (per i tipi tabellari e discrezionali)"
COL_CONFIDENCE = "Livello di confidenza dell'analisi"
COL_MOTIVATION = (
    "Motivazioni e argomentazioni a supporto del punteggio assegnato "
    "(per i tabellari e discrezionali)"
)
COL_DOCS = "Nome del/dei documenti utilizzati ove è trattato il criterio"
COL_PAGES = "n. pagina/e del/dei documento/i ove è trattato il criterio"
COL_NOTES = "Note e/o descrizione aggiuntiva a supporto dell'analisi condotta"

HEADERS = [
    COL_OPERATOR, COL_CRITERIO, COL_TIPO, COL_MODE, COL_MAX_POINTS,
    COL_PRESENZA, COL_SCORE, COL_CONFIDENCE, COL_MOTIVATION, COL_DOCS, COL_PAGES, COL_NOTES,
]

_HUMAN_REVIEW_MARK = "⚠ REVISIONE UMANA"
_CITATION_UNATTRIBUTED = "⚠ da verificare"

_HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


class RestituzioneXlsxService:
    """Build the restituzione (output) workbook from one or more check reports."""

    def build_workbook(
        self,
        reports: List[ComplianceReportV2],
        operator_labels: Optional[Dict[str, str]] = None,
    ) -> bytes:
        """Render *reports* into an xlsx. One sheet per lot; rows per (criterion, operator).

        *operator_labels* maps namespace → display label; falls back to the namespace.
        """
        operator_labels = operator_labels or {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Group reports per lot (preserve first-seen order).
        lots: Dict[str, List[ComplianceReportV2]] = {}
        for rep in reports:
            lots.setdefault(rep.tender.lot_id, []).append(rep)

        used: set = set()
        for lot_id, lot_reports in lots.items():
            title = self._sheet_title(lot_reports[0].tender.lot_name or lot_id, used)
            ws = wb.create_sheet(title=title)
            self._write_lot_sheet(ws, lot_reports, operator_labels)

        if not wb.worksheets:
            wb.create_sheet(title="Restituzione")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _sheet_title(self, base: str, used: set) -> str:
        base = str(base).strip() or "Lotto"
        for ch in r"[]:*?/\\":
            base = base.replace(ch, " ")
        base = base[:28].strip() or "Lotto"
        title = base
        i = 2
        while title in used:
            suffix = f" {i}"
            title = base[:28 - len(suffix)] + suffix
            i += 1
        used.add(title)
        return title

    def _label(self, rep: ComplianceReportV2, labels: Dict[str, str]) -> str:
        return labels.get(rep.namespace) or rep.namespace

    def _write_lot_sheet(
        self,
        ws,
        reports: List[ComplianceReportV2],
        labels: Dict[str, str],
    ) -> None:
        header_row = 1
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _WRAP
        row = header_row + 1

        # Criterion order is taken from the first report (criteria are shared per lot).
        first = reports[0]
        tab_ids = [r.requirement_id for r in first.tabular_results]
        disc_ids = [d.criterion_id for d in first.discretionary_results]

        tab_by_op = {
            self._label(rep, labels): {r.requirement_id: r for r in rep.tabular_results}
            for rep in reports
        }
        disc_by_op = {
            self._label(rep, labels): {d.criterion_id: d for d in rep.discretionary_results}
            for rep in reports
        }
        operators = [self._label(rep, labels) for rep in reports]

        # Conformità (tabular) first, then scored criteria — by criterion, then operator.
        for rid in tab_ids:
            for op in operators:
                r = tab_by_op.get(op, {}).get(rid)
                if r is None:
                    continue
                row = self._write_tabular_row(ws, row, op, r)

        for cid in disc_ids:
            for op in operators:
                d = disc_by_op.get(op, {}).get(cid)
                if d is None:
                    continue
                row = self._write_discretionary_row(ws, row, op, d)

        self._apply_formatting(ws, header_row, last_row=row - 1)

    def _write_tabular_row(self, ws, row: int, operator: str, r) -> int:
        presenza = _DEFAULT_JUDGMENT_MAP.get(r.judgment, r.judgment)
        values = [
            operator,
            r.requirement_text,
            tax.TYPE_CONFORMITA,
            "",                       # Modalità (n/a for Conformità)
            "",                       # Punteggio previsto (n/a)
            presenza,
            "",                       # Punteggio assegnato (n/a for Conformità)
            round(r.confidence, 2),
            r.justification,
            r.evidence_document,
            str(r.evidence_page) if r.evidence_document else "",
            "",
        ]
        self._emit(ws, row, values)
        return row + 1

    def _write_discretionary_row(self, ws, row: int, operator: str, d) -> int:
        if d.human_review_required:
            if d.proportional_auto and d.score is not None:
                score_cell = f"{round(d.score, 2)} (da confermare)"
            else:
                score_cell = _HUMAN_REVIEW_MARK
        elif d.score is not None:
            score_cell = round(d.score, 2)
        elif d.measured_value:
            score_cell = f"Valore: {d.measured_value}"
        else:
            score_cell = "N/V"

        if not d.citation_attributed:
            docs = _CITATION_UNATTRIBUTED
            pages = ""
        else:
            docs = d.evidence_document
            pages = str(d.evidence_page) if d.evidence_document else ""

        note_bits = []
        if d.human_review_reason:
            note_bits.append(d.human_review_reason)
        if not d.citation_attributed:
            note_bits.append(
                "Citazione non attribuita a un chunk: riferimento da verificare."
            )

        values = [
            operator,
            d.criterion_text,
            tax.type_from_mode(d.mode),
            d.mode.value,
            d.max_points,
            "",                       # Presenza (n/a for scored criteria)
            score_cell,
            round(d.confidence, 2),
            d.motivation,
            docs,
            pages,
            " ".join(note_bits),
        ]
        self._emit(ws, row, values)
        return row + 1

    def _emit(self, ws, row: int, values: List) -> None:
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = _WRAP

    def _apply_formatting(self, ws, header_row: int, last_row: int) -> None:
        widths = [22, 60, 14, 14, 12, 14, 16, 14, 50, 28, 12, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
