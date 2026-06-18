"""
RequirementsXlsxService — round-trip between TenderLotRequirements and the
*standardized, human-reviewable* criteria workbook (PA / market template).

This is NOT the free-form capitolato parsed by XlsxExtractionService (LLM-based).
It is a deterministic, structured workbook aligned to `ipotesi tabella criteri.xlsx`:

  - `build_workbook(lots)`  → one sheet per lot, business taxonomy
    (Conformità / Tabellare / Discrezionale), ready for review in Excel.
  - `parse_workbook(bytes)` → reads it back into TenderLotRequirements WITHOUT any
    LLM, so the reviewed file can be fed straight into /v2/check.

Sheet layout (criteria are operator-agnostic — no "Operatore Economico" column here;
that column belongs to the restituzione/output workbook):

    Lotto                | <lot_id>
    Nome lotto           | <lot_name>
    Titolo               | <title>
    File sorgente        | <source_file>
    Avvisi di revisione  | (optional warning rows)
    <blank>
    Criterio | Tipo criterio | Modalità | Punteggio previsto | Note... | Obbligatorio | Solo revisione umana
    ...data rows...

`Tipo criterio` uses the business taxonomy (see xlsx_taxonomy). `Modalità`,
`Obbligatorio` and `Solo revisione umana` are tolerated-if-absent extra columns: a
file carrying only the core market columns still parses — the mode is then derived
from the criterion text and the flags fall back to their defaults.
"""
import io
import logging
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from tilellm.modules.compliance_checker.models_v2 import (
    DiscretionaryCriterion,
    DiscretionaryMode,
    TenderInfo,
    TenderLotRequirements,
    TabularRequirementV2,
    _RequirementsBlock,
)
from tilellm.modules.compliance_checker.services import xlsx_taxonomy as tax

logger = logging.getLogger(__name__)

MAX_XLSX_SIZE: int = 10 * 1024 * 1024  # 10 MB

# --- Metadata labels (column A) --------------------------------------------
_LBL_LOT_ID = "Lotto"
_LBL_LOT_NAME = "Nome lotto"
_LBL_TITLE = "Titolo"
_LBL_SOURCE_FILE = "File sorgente"
_LBL_WARNINGS = "Avvisi di revisione"
_META_LABELS = {_LBL_LOT_ID, _LBL_LOT_NAME, _LBL_TITLE, _LBL_SOURCE_FILE}

# --- Table headers (criteria template) -------------------------------------
_COL_CRITERIO = "Criterio"
_COL_TIPO = "Tipo criterio"
_COL_MODE = "Modalità"
_COL_MAX_POINTS = "Punteggio previsto"
_COL_NOTES = "Note e/o descrizione aggiuntiva"
_COL_MANDATORY = "Obbligatorio"
_COL_HUMAN_ONLY = "Solo revisione umana"
_COL_ID = "ID"  # optional — honored if present, else auto-generated

_HEADERS = [
    _COL_CRITERIO, _COL_TIPO, _COL_MODE, _COL_MAX_POINTS,
    _COL_NOTES, _COL_MANDATORY, _COL_HUMAN_ONLY, _COL_ID,
]

_HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_META_FONT = Font(bold=True)
_VALID_MODES = ",".join(sorted(m.value for m in DiscretionaryMode))


class RequirementsXlsxService:
    """Build / parse the standardized criteria workbook (no LLM involved)."""

    # ------------------------------------------------------------------
    # Build
    # ------------------------------------------------------------------

    def build_workbook(
        self,
        lots: List[TenderLotRequirements],
        warnings_by_lot: Optional[Dict[str, List[str]]] = None,
    ) -> bytes:
        """Serialize *lots* into an xlsx workbook (one sheet per lot)."""
        warnings_by_lot = warnings_by_lot or {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        used: set = set()
        for lot in lots:
            ws = wb.create_sheet(title=self._sheet_title(lot, used))
            self._write_lot_sheet(ws, lot, warnings_by_lot.get(lot.tender.lot_id, []))

        if not wb.worksheets:
            wb.create_sheet(title="Requisiti")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _sheet_title(self, lot: TenderLotRequirements, used: set) -> str:
        base = (lot.tender.lot_name or lot.tender.lot_id or "Lotto").strip()
        for ch in r"[]:*?/\\":
            base = base.replace(ch, " ")
        base = base[:28].strip() or "Lotto"
        title = base
        i = 2
        while title in used:
            suffix = f" {i}"
            title = base[:28 - len(suffix)] + suffix
            i += 1
        used.add(title)
        return title

    def _write_lot_sheet(self, ws, lot: TenderLotRequirements, warnings: List[str]) -> None:
        t = lot.tender
        row = 1
        for label, value in [
            (_LBL_LOT_ID, t.lot_id),
            (_LBL_LOT_NAME, t.lot_name),
            (_LBL_TITLE, t.title),
            (_LBL_SOURCE_FILE, t.source_file or ""),
        ]:
            ws.cell(row=row, column=1, value=label).font = _META_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1

        if warnings:
            ws.cell(row=row, column=1, value=_LBL_WARNINGS).font = _META_FONT
            for w in warnings:
                ws.cell(row=row, column=2, value=f"⚠️ {w}")
                row += 1

        row += 1  # blank separator

        header_row = row
        for col, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = _META_FONT
            cell.fill = _HEADER_FILL
        row += 1

        for r in lot.requirements.tabular:
            ws.cell(row=row, column=1, value=r.text)
            ws.cell(row=row, column=2, value=tax.TYPE_CONFORMITA)
            ws.cell(row=row, column=4, value=0)
            ws.cell(row=row, column=6, value=tax.bool_to_cell(r.mandatory))
            ws.cell(row=row, column=8, value=r.id)
            row += 1

        for c in lot.requirements.discretionary:
            ws.cell(row=row, column=1, value=c.text)
            ws.cell(row=row, column=2, value=tax.model_type(c))
            ws.cell(row=row, column=3, value=tax.model_mode_label(c))
            ws.cell(row=row, column=4, value=c.max_points)
            ws.cell(row=row, column=5, value=c.notes or "")
            ws.cell(row=row, column=7, value=tax.bool_to_cell(c.human_only))
            ws.cell(row=row, column=8, value=c.id)
            row += 1

        self._apply_formatting(ws, header_row, last_row=row - 1)

    def _apply_formatting(self, ws, header_row: int, last_row: int) -> None:
        widths = [60, 16, 16, 18, 40, 14, 20, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        if last_row < header_row + 1:
            return
        first = header_row + 1
        dv_type = DataValidation(
            type="list",
            formula1=f'"{tax.TYPE_CONFORMITA},{tax.TYPE_TABELLARE},{tax.TYPE_DISCREZIONALE}"',
            allow_blank=True,
        )
        ws.add_data_validation(dv_type)
        dv_type.add(f"B{first}:B{last_row}")
        dv_mode = DataValidation(type="list", formula1=f'"{_VALID_MODES}"', allow_blank=True)
        ws.add_data_validation(dv_mode)
        dv_mode.add(f"C{first}:C{last_row}")
        dv_bool = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
        ws.add_data_validation(dv_bool)
        dv_bool.add(f"F{first}:F{last_row}")
        dv_bool.add(f"G{first}:G{last_row}")

    # ------------------------------------------------------------------
    # Parse
    # ------------------------------------------------------------------

    def parse_workbook(self, xlsx_bytes: bytes) -> List[TenderLotRequirements]:
        """Read a standardized criteria workbook back into TenderLotRequirements (no LLM)."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"File xlsx non leggibile: {e}") from e

        lots: List[TenderLotRequirements] = []
        for ws in wb.worksheets:
            lot = self._parse_sheet(ws)
            if lot is not None:
                lots.append(lot)
        wb.close()

        if not lots:
            raise ValueError(
                "Nessun lotto valido trovato nel file xlsx: assicurarsi che ogni foglio "
                f"contenga la riga di intestazione che inizia con '{_COL_CRITERIO}'."
            )
        return lots

    def _parse_sheet(self, ws) -> Optional[TenderLotRequirements]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        metadata: Dict[str, str] = {}
        header_idx: Optional[int] = None
        col_map: Dict[str, int] = {}

        for idx, raw in enumerate(rows):
            cells = list(raw)
            first = (str(cells[0]).strip() if cells and cells[0] is not None else "")
            if first == _COL_CRITERIO:
                header_idx = idx
                for col, value in enumerate(cells):
                    key = (str(value).strip() if value is not None else "")
                    if key:
                        col_map[key] = col
                break
            if first in _META_LABELS:
                value = cells[1] if len(cells) > 1 and cells[1] is not None else ""
                metadata[first] = str(value).strip()

        if header_idx is None:
            logger.debug("Sheet '%s' has no '%s' header — skipped.", ws.title, _COL_CRITERIO)
            return None

        lot_id = metadata.get(_LBL_LOT_ID) or ws.title
        lot_name = metadata.get(_LBL_LOT_NAME) or ws.title
        title = metadata.get(_LBL_TITLE) or lot_name
        source_file = metadata.get(_LBL_SOURCE_FILE) or None

        def _get(cells, header: str):
            col = col_map.get(header)
            if col is None or col >= len(cells):
                return None
            return cells[col]

        tabular: List[TabularRequirementV2] = []
        discretionary: List[DiscretionaryCriterion] = []
        counters = {tax.TYPE_CONFORMITA: 0, tax.TYPE_TABELLARE: 0, tax.TYPE_DISCREZIONALE: 0}

        for raw in rows[header_idx + 1:]:
            cells = list(raw)
            text = _get(cells, _COL_CRITERIO)
            text = str(text).strip() if text is not None else ""
            tipo = tax.normalize_type(_get(cells, _COL_TIPO))
            if not text and tipo is None:
                continue  # blank / trailing row
            if tipo is None:
                logger.warning(
                    "Riga senza 'Tipo criterio' riconoscibile in '%s' — ignorata: %.60s",
                    ws.title, text,
                )
                continue

            counters[tipo] += 1
            explicit_id = _get(cells, _COL_ID)
            explicit_id = str(explicit_id).strip() if explicit_id not in (None, "") else ""

            if tipo == tax.TYPE_CONFORMITA:
                rid = explicit_id or f"C{counters[tipo]}"
                mandatory = tax.cell_to_bool(_get(cells, _COL_MANDATORY), default=True)
                tabular.append(TabularRequirementV2(id=rid, text=text, mandatory=mandatory))
            else:
                prefix = "T" if tipo == tax.TYPE_TABELLARE else "D"
                rid = explicit_id or f"{prefix}{counters[tipo]}"
                default_mode = (
                    DiscretionaryMode.VARIABILE
                    if tipo == tax.TYPE_DISCREZIONALE
                    else DiscretionaryMode.ON_OFF
                )
                mode, warn = tax.resolve_mode(
                    _get(cells, _COL_MODE), text, default=default_mode
                )
                if warn:
                    logger.warning("Criterio '%s' (%s): %s", rid, ws.title, warn)
                max_points = tax.parse_float(_get(cells, _COL_MAX_POINTS), row_id=rid)
                human_only = tax.cell_to_bool(_get(cells, _COL_HUMAN_ONLY), default=False)
                notes = _get(cells, _COL_NOTES)
                notes = str(notes).strip() if notes not in (None, "") else None
                discretionary.append(
                    DiscretionaryCriterion(
                        id=rid, text=text, mode=mode, max_points=max_points,
                        human_only=human_only, notes=notes,
                    )
                )

        return TenderLotRequirements(
            tender=TenderInfo(
                title=title, lot_id=lot_id, lot_name=lot_name, source_file=source_file,
            ),
            requirements=_RequirementsBlock(tabular=tabular, discretionary=discretionary),
        )


def select_lot(
    lots: List[TenderLotRequirements],
    lot_id: Optional[str],
) -> TenderLotRequirements:
    """Pick a single lot from the parsed workbook."""
    if not lots:
        raise ValueError("Nessun lotto trovato nel file xlsx.")
    if lot_id:
        for lot in lots:
            if lot.tender.lot_id == lot_id:
                return lot
        available = ", ".join(repr(l.tender.lot_id) for l in lots)
        raise ValueError(
            f"lot_id '{lot_id}' non trovato nel file xlsx. Lotti disponibili: {available}."
        )
    if len(lots) == 1:
        return lots[0]
    available = ", ".join(repr(l.tender.lot_id) for l in lots)
    raise ValueError(
        f"Il file xlsx contiene più lotti ({available}): specificare 'requirements_lot_id' "
        f"per indicare quale valutare."
    )
