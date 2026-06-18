"""
Unit tests for ComplianceChecker v2 — FASE 1 (models) + FASE 3/5 (services).

Coverage target: ≥ 80% su models_v2.py e services/
Follows AAA pattern (Arrange / Act / Assert).
LLM and repo dependencies are mocked via unittest.mock.
"""
import io
import json
import pytest
import yaml

from tilellm.modules.compliance_checker.models_v2 import (
    AggregateReport,
    AggregateReportRequest,
    ComplianceReportV2,
    ComplianceRequestV2,
    ComplianceSummaryV2,
    DiscretionaryCriterion,
    DiscretionaryMode,
    DiscretionaryResult,
    ExtractRequirementsRequest,
    ExtractRequirementsResponse,
    HumanReviewItem,
    LotComparison,
    LotYamlDocument,
    OperatorScore,
    ReportSource,
    TabularRequirementV2,
    TabularSummary,
    TenderInfo,
    TenderLotRequirements,
)
from tilellm.modules.compliance_checker.models import ComplianceResult


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

MINIMAL_YAML = """\
tender:
  title: "Gara test"
  lot_id: "6"
  lot_name: "Lotto 6"
  source_file: "test.xlsx"
requirements:
  tabular:
    - id: REQ-001
      text: "sterile"
      mandatory: true
  discretionary:
    - id: P1
      text: "plasticità del tubo"
      mode: variabile
      max_points: 8
      human_only: false
"""

MULTI_LOT_YAML = """\
tender:
  title: "Gara test multi-lotto"
  lot_id: "9"
  lot_name: "Lotto 9"
  source_file: "test.xlsx"
requirements:
  tabular: []
  discretionary:
    - id: P1
      text: "morbidezza cuffia"
      mode: variabile
      max_points: 12
    - id: P2
      text: "ampiezza gamma"
      mode: proporzionale
      max_points: 10
    - id: P3
      text: "certificato UNI PDR"
      mode: on_off
      max_points: 1
    - id: P4
      text: "ergonomicità impugnatura"
      mode: variabile
      max_points: 11
      human_only: true
      notes: "Criterio soggettivo, non gestibile da IA"
"""


def _make_discretionary_result(
    criterion_id: str = "P1",
    mode: DiscretionaryMode = DiscretionaryMode.VARIABILE,
    max_points: float = 8.0,
    coefficient: float | None = 0.75,
    score: float | None = 6.0,
    human_review_required: bool = False,
    human_review_reason: str | None = None,
    measured_value: str | None = None,
    citation_attributed: bool = True,
) -> DiscretionaryResult:
    return DiscretionaryResult(
        criterion_id=criterion_id,
        criterion_text="testo criterio",
        mode=mode,
        max_points=max_points,
        coefficient=coefficient,
        score=score,
        measured_value=measured_value,
        motivation="Buona documentazione.",
        confidence=0.85,
        human_review_required=human_review_required,
        human_review_reason=human_review_reason,
        citation_attributed=citation_attributed,
        evidence_document="offerta.pdf",
        evidence_page=5,
        evidence_section="Cap. 3",
        evidence_text="testo citato",
        evidence_chunk_index=1,
        evidence_chunk_ids=["id1"],
    )


def _make_tabular_result(req_id: str = "REQ-001", judgment: str = "compliant") -> ComplianceResult:
    return ComplianceResult(
        requirement_id=req_id,
        requirement_text="sterile",
        category=None,
        mandatory=True,
        judgment=judgment,
        confidence=0.9,
        evidence_text="il prodotto è sterile",
        justification="Conforme.",
        evidence_document="offerta.pdf",
        evidence_page=1,
        evidence_section="",
        evidence_chunk_index=1,
        evidence_chunk_ids=[],
    )


# ---------------------------------------------------------------------------
# TASK 1.1 — DiscretionaryMode
# ---------------------------------------------------------------------------

class TestDiscretionaryMode:

    def test_variabile_value(self):
        assert DiscretionaryMode.VARIABILE == "variabile"

    def test_proporzionale_value(self):
        assert DiscretionaryMode.PROPORZIONALE == "proporzionale"

    def test_on_off_value(self):
        assert DiscretionaryMode.ON_OFF == "on_off"

    def test_is_string(self):
        assert isinstance(DiscretionaryMode.VARIABILE, str)

    def test_invalid_mode_raises(self):
        with pytest.raises(Exception):
            DiscretionaryCriterion(id="P1", text="x", mode="invalid_mode", max_points=5)


# ---------------------------------------------------------------------------
# TASK 1.2 — TabularRequirementV2
# ---------------------------------------------------------------------------

class TestTabularRequirementV2:

    def test_creation_minimal(self):
        req = TabularRequirementV2(id="REQ-001", text="sterile")
        assert req.id == "REQ-001"
        assert req.mandatory is True

    def test_mandatory_default_true(self):
        req = TabularRequirementV2(id="R", text="t")
        assert req.mandatory is True

    def test_optional_mandatory_false(self):
        req = TabularRequirementV2(id="R", text="t", mandatory=False)
        assert req.mandatory is False


# ---------------------------------------------------------------------------
# TASK 1.3 — DiscretionaryCriterion
# ---------------------------------------------------------------------------

class TestDiscretionaryCriterion:

    def test_creation_valid(self):
        c = DiscretionaryCriterion(id="P1", text="plasticità", mode="variabile", max_points=8)
        assert c.mode == DiscretionaryMode.VARIABILE
        assert c.max_points == 8.0
        assert c.human_only is False

    def test_max_points_zero_raises(self):
        with pytest.raises(Exception):
            DiscretionaryCriterion(id="P1", text="x", mode="variabile", max_points=0)

    def test_max_points_negative_raises(self):
        with pytest.raises(Exception):
            DiscretionaryCriterion(id="P1", text="x", mode="variabile", max_points=-1)

    def test_human_only_default_false(self):
        c = DiscretionaryCriterion(id="P1", text="x", mode="on_off", max_points=1)
        assert c.human_only is False

    def test_proporzionale_mode(self):
        c = DiscretionaryCriterion(id="P5", text="ampiezza gamma", mode="proporzionale", max_points=5)
        assert c.mode == DiscretionaryMode.PROPORZIONALE

    def test_on_off_mode(self):
        c = DiscretionaryCriterion(id="P9", text="cert UNI", mode="on_off", max_points=1)
        assert c.mode == DiscretionaryMode.ON_OFF


# ---------------------------------------------------------------------------
# TASK 1.4 — TenderInfo
# ---------------------------------------------------------------------------

class TestTenderInfo:

    def test_creation(self):
        t = TenderInfo(title="Gara", lot_id="6", lot_name="Lotto 6")
        assert t.lot_id == "6"
        assert t.source_file is None

    def test_optional_fields(self):
        t = TenderInfo(title="G", lot_id="1", lot_name="L1", source_file="file.xlsx")
        assert t.source_file == "file.xlsx"


# ---------------------------------------------------------------------------
# TASK 1.5 — TenderLotRequirements
# ---------------------------------------------------------------------------

class TestTenderLotRequirements:

    def test_from_yaml_minimal(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        assert lot.tender.lot_id == "6"
        assert len(lot.requirements.tabular) == 1
        assert len(lot.requirements.discretionary) == 1

    def test_to_yaml_roundtrip(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        yaml_out = lot.to_yaml()
        lot2 = TenderLotRequirements.from_yaml(yaml_out)
        assert lot2.tender.lot_id == lot.tender.lot_id
        assert lot2.requirements.discretionary[0].mode == lot.requirements.discretionary[0].mode

    def test_from_yaml_invalid_yaml_raises_value_error(self):
        with pytest.raises(ValueError, match="YAML"):
            TenderLotRequirements.from_yaml(":::invalid::yaml:::")

    def test_from_yaml_invalid_mode_raises(self):
        bad = MINIMAL_YAML.replace("mode: variabile", "mode: sconosciuto")
        with pytest.raises(Exception):
            TenderLotRequirements.from_yaml(bad)

    def test_from_yaml_max_points_zero_raises(self):
        bad = MINIMAL_YAML.replace("max_points: 8", "max_points: 0")
        with pytest.raises(Exception):
            TenderLotRequirements.from_yaml(bad)

    def test_from_yaml_duplicate_discretionary_ids_raises(self):
        bad = MINIMAL_YAML + "    - id: P1\n      text: duplicato\n      mode: variabile\n      max_points: 5\n"
        with pytest.raises(ValueError, match="duplicat"):
            TenderLotRequirements.from_yaml(bad)

    def test_from_yaml_duplicate_tabular_ids_raises(self):
        bad = """\
tender:
  title: "Gara test"
  lot_id: "6"
  lot_name: "Lotto 6"
requirements:
  tabular:
    - id: REQ-001
      text: "sterile"
      mandatory: true
    - id: REQ-001
      text: "duplicato"
      mandatory: true
  discretionary: []
"""
        with pytest.raises(ValueError, match="duplicat"):
            TenderLotRequirements.from_yaml(bad)

    def test_from_yaml_multi_lot(self):
        lot = TenderLotRequirements.from_yaml(MULTI_LOT_YAML)
        assert lot.tender.lot_id == "9"
        assert len(lot.requirements.discretionary) == 4
        human_only = [c for c in lot.requirements.discretionary if c.human_only]
        assert len(human_only) == 1
        assert human_only[0].id == "P4"

    def test_from_yaml_empty_lists(self):
        minimal_no_disc = MINIMAL_YAML.replace(
            "  discretionary:\n    - id: P1\n      text: \"plasticità del tubo\"\n      mode: variabile\n      max_points: 8\n      human_only: false\n",
            "  discretionary: []\n"
        )
        lot = TenderLotRequirements.from_yaml(minimal_no_disc)
        assert lot.requirements.discretionary == []

    def test_to_yaml_produces_valid_yaml(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        yaml_text = lot.to_yaml()
        parsed = yaml.safe_load(yaml_text)
        assert "tender" in parsed
        assert "requirements" in parsed

    def test_to_yaml_unicode_preserved(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        lot.requirements.discretionary[0].text = "plasticità del tubo (caratteri speciali: àèìòù)"
        yaml_text = lot.to_yaml()
        assert "plasticità" in yaml_text


# ---------------------------------------------------------------------------
# TASK 1.6 — ExtractRequirementsRequest
# ---------------------------------------------------------------------------

class TestExtractRequirementsRequest:

    def test_creation_minimal(self):
        req = ExtractRequirementsRequest(source="https://example.com/file.xlsx")
        assert req.source == "https://example.com/file.xlsx"
        assert req.model == "gpt-4o-mini"

    def test_llm_fields_defaults(self):
        req = ExtractRequirementsRequest(source="https://x.com/f.xlsx")
        assert req.llm == "openai"
        assert req.temperature == 0.0


# ---------------------------------------------------------------------------
# TASK 1.7 — ExtractRequirementsResponse / LotYamlDocument
# ---------------------------------------------------------------------------

class TestExtractRequirementsResponse:

    def test_lot_yaml_document(self):
        doc = LotYamlDocument(lot_id="6", lot_name="Lotto 6", yaml="tender:\n  lot_id: '6'")
        assert doc.lot_id == "6"
        assert doc.warnings == []

    def test_response_with_warnings(self):
        doc = LotYamlDocument(
            lot_id="6",
            lot_name="Lotto 6",
            yaml="tender:\n  lot_id: '6'",
            warnings=["P8: human_only"],
        )
        resp = ExtractRequirementsResponse(lots=[doc])
        assert len(resp.lots) == 1
        assert resp.lots[0].warnings == ["P8: human_only"]

    def test_response_empty_lots(self):
        resp = ExtractRequirementsResponse(lots=[])
        assert resp.lots == []


# ---------------------------------------------------------------------------
# TASK 1.8 — ComplianceRequestV2
# ---------------------------------------------------------------------------

class TestComplianceRequestV2:

    _ENGINE = {
        "name": "pinecone",
        "type": "serverless",
        "apikey": "key",
        "vector_size": 1536,
        "index_name": "idx",
    }

    def test_yaml_inline_only(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML,
            namespace="ns",
            engine=self._ENGINE,
        )
        assert req.requirements_yaml == MINIMAL_YAML
        assert req.requirements_yaml_url is None

    def test_yaml_url_only(self):
        req = ComplianceRequestV2(
            requirements_yaml_url="https://example.com/lot6.yaml",
            namespace="ns",
            engine=self._ENGINE,
        )
        assert req.requirements_yaml_url == "https://example.com/lot6.yaml"
        assert req.requirements_yaml is None

    def test_both_raises(self):
        with pytest.raises(ValueError, match="mutuamente esclusivi"):
            ComplianceRequestV2(
                requirements_yaml=MINIMAL_YAML,
                requirements_yaml_url="https://x.com/f.yaml",
                namespace="ns",
                engine=self._ENGINE,
            )

    def test_neither_raises(self):
        with pytest.raises(ValueError, match="requirements_yaml"):
            ComplianceRequestV2(namespace="ns", engine=self._ENGINE)

    def test_min_confidence_default(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML,
            namespace="ns",
            engine=self._ENGINE,
        )
        assert req.min_confidence == 0.6

    def test_min_confidence_custom(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML,
            namespace="ns",
            engine=self._ENGINE,
            min_confidence=0.75,
        )
        assert req.min_confidence == 0.75

    def test_llm_defaults(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML,
            namespace="ns",
            engine=self._ENGINE,
        )
        assert req.llm == "openai"
        assert req.model == "gpt-4o-mini"
        assert req.top_k == 8

    def test_reranker_config_disabled_by_default(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML, namespace="ns", engine=self._ENGINE,
        )
        assert req.reranking is False
        assert req.reranker_config is None

    def test_reranker_config_true_resolves_to_model(self):
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML, namespace="ns", engine=self._ENGINE,
            reranking=True, reranker_model="cross-encoder/ms-marco-MiniLM-L-6-v2",
        )
        assert req.reranker_config == "cross-encoder/ms-marco-MiniLM-L-6-v2"

    def test_reranker_config_object_passthrough(self):
        from tilellm.models.llm import TEIConfig
        tei = TEIConfig(name="BAAI/bge-reranker-base", url="http://tei:8080")
        req = ComplianceRequestV2(
            requirements_yaml=MINIMAL_YAML, namespace="ns", engine=self._ENGINE,
            reranking=tei,
        )
        assert req.reranker_config is tei


# ---------------------------------------------------------------------------
# TASK 1.9 — DiscretionaryResult
# ---------------------------------------------------------------------------

class TestDiscretionaryResult:

    def test_variabile_result(self):
        r = _make_discretionary_result(coefficient=0.75, score=6.0)
        assert r.coefficient == 0.75
        assert r.score == 6.0
        assert r.human_review_required is False

    def test_proporzionale_result(self):
        r = _make_discretionary_result(
            mode=DiscretionaryMode.PROPORZIONALE,
            coefficient=None,
            score=None,
            measured_value="14 misure 4.0–10.5 mm",
            human_review_required=True,
            human_review_reason="Confronto comparativo tra operatori richiesto.",
        )
        assert r.coefficient is None
        assert r.score is None
        assert r.measured_value == "14 misure 4.0–10.5 mm"
        assert r.human_review_required is True

    def test_human_only_result(self):
        r = _make_discretionary_result(
            coefficient=None,
            score=None,
            human_review_required=True,
            human_review_reason="Criterio soggettivo, non gestibile da IA.",
        )
        assert r.human_review_reason is not None

    def test_on_off_result_full_score(self):
        r = _make_discretionary_result(
            mode=DiscretionaryMode.ON_OFF,
            coefficient=1.0,
            score=1.0,
            max_points=1.0,
        )
        assert r.score == 1.0

    def test_on_off_result_zero_score(self):
        r = _make_discretionary_result(
            mode=DiscretionaryMode.ON_OFF,
            coefficient=0.0,
            score=0.0,
            max_points=1.0,
        )
        assert r.score == 0.0

    def test_citation_attributed_defaults_true(self):
        r = DiscretionaryResult(
            criterion_id="P1", criterion_text="x", mode=DiscretionaryMode.VARIABILE,
            max_points=8, motivation="m", confidence=0.8,
        )
        assert r.citation_attributed is True


# ---------------------------------------------------------------------------
# TASK 1.10 — ComplianceSummaryV2
# ---------------------------------------------------------------------------

class TestComplianceSummaryV2:

    def test_defaults(self):
        s = ComplianceSummaryV2()
        assert s.tabular.total == 0
        assert s.discretionary_total == 0
        assert s.ai_scored_points == 0.0
        assert s.human_review_count == 0

    def test_tabular_summary_compliance_rate(self):
        ts = TabularSummary(total=5, compliant=4, non_compliant=1, not_verifiable=0)
        assert ts.compliance_rate == pytest.approx(0.8, abs=1e-3)

    def test_tabular_summary_all_not_verifiable(self):
        ts = TabularSummary(total=3, not_verifiable=3)
        assert ts.compliance_rate == 0.0

    def test_summary_from_results(self):
        tabular_results = [
            _make_tabular_result("R1", "compliant"),
            _make_tabular_result("R2", "non_compliant"),
            _make_tabular_result("R3", "not_verifiable"),
        ]
        disc_results = [
            _make_discretionary_result("P1", score=6.0, coefficient=0.75),
            _make_discretionary_result(
                "P2",
                mode=DiscretionaryMode.PROPORZIONALE,
                coefficient=None,
                score=None,
                human_review_required=True,
                human_review_reason="proporzionale",
            ),
        ]
        summary = ComplianceSummaryV2.from_results(tabular_results, disc_results)
        assert summary.tabular.total == 3
        assert summary.tabular.compliant == 1
        assert summary.discretionary_total == 2
        assert summary.ai_scored_count == 1
        assert summary.ai_scored_points == pytest.approx(6.0)
        assert summary.human_review_count == 1

    def test_summary_counts_unattributed_citations(self):
        disc_results = [
            _make_discretionary_result("P1", score=6.0, coefficient=0.75, citation_attributed=True),
            _make_discretionary_result("P2", score=4.0, coefficient=0.5, citation_attributed=False),
            _make_discretionary_result("P3", score=2.0, coefficient=0.25, citation_attributed=False),
        ]
        summary = ComplianceSummaryV2.from_results([], disc_results)
        assert summary.citation_unattributed_count == 2


# ---------------------------------------------------------------------------
# TASK 1.11 — ComplianceReportV2
# ---------------------------------------------------------------------------

class TestComplianceReportV2:

    def _make_report(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        tabular = [_make_tabular_result("REQ-001", "compliant")]
        disc = [_make_discretionary_result("P1", score=6.0, coefficient=0.75)]
        summary = ComplianceSummaryV2.from_results(tabular, disc)
        return ComplianceReportV2(
            tender=lot.tender,
            namespace="ns-operatore-1",
            summary=summary,
            tabular_results=tabular,
            discretionary_results=disc,
        )

    def test_creation(self):
        report = self._make_report()
        assert report.namespace == "ns-operatore-1"
        assert report.tender.lot_id == "6"

    def test_to_markdown_contains_headers(self):
        report = self._make_report()
        md = report.to_markdown()
        assert "Requisito" in md
        assert "Punteggio" in md
        assert "Documento" in md
        assert "Pagina" in md
        assert "Motivazione" in md

    def test_to_markdown_shows_citation_audit_marker(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        disc = [_make_discretionary_result("P1", score=6.0, coefficient=0.75, citation_attributed=False)]
        summary = ComplianceSummaryV2.from_results([], disc)
        report = ComplianceReportV2(
            tender=lot.tender, namespace="ns", summary=summary,
            tabular_results=[], discretionary_results=disc,
        )
        md = report.to_markdown()
        assert "citazione non attribuita" in md.lower()

    def test_to_markdown_contains_tabular_result(self):
        report = self._make_report()
        md = report.to_markdown()
        assert "sterile" in md
        assert "SI" in md

    def test_to_markdown_contains_discretionary_result(self):
        report = self._make_report()
        md = report.to_markdown()
        assert "P1" in md or "testo criterio" in md
        assert "6.0" in md or "6" in md

    def test_to_markdown_human_review_flagged(self):
        lot = TenderLotRequirements.from_yaml(MINIMAL_YAML)
        disc = [
            _make_discretionary_result(
                "P1",
                coefficient=None,
                score=None,
                human_review_required=True,
                human_review_reason="Criterio soggettivo.",
            )
        ]
        summary = ComplianceSummaryV2.from_results([], disc)
        report = ComplianceReportV2(
            tender=lot.tender,
            namespace="ns",
            summary=summary,
            tabular_results=[],
            discretionary_results=disc,
        )
        md = report.to_markdown()
        assert "⚠" in md or "UMANO" in md or "human" in md.lower() or "revisione" in md.lower()


# ===========================================================================
# FASE 3/5 — Services
# ===========================================================================

from unittest.mock import AsyncMock, MagicMock, Mock, patch


# ---------------------------------------------------------------------------
# TASK 5.2 — YamlRequirementsLoader
# ---------------------------------------------------------------------------

class TestYamlRequirementsLoader:

    @pytest.mark.asyncio
    async def test_load_inline(self):
        from tilellm.modules.compliance_checker.services.yaml_requirements_loader import YamlRequirementsLoader
        loader = YamlRequirementsLoader()
        lot = await loader.load(yaml_inline=MINIMAL_YAML, yaml_url=None)
        assert lot.tender.lot_id == "6"
        assert len(lot.requirements.tabular) == 1

    @pytest.mark.asyncio
    async def test_load_url(self):
        from tilellm.modules.compliance_checker.services.yaml_requirements_loader import YamlRequirementsLoader
        loader = YamlRequirementsLoader()
        mock_response = MagicMock()
        mock_response.raise_for_status = Mock()
        mock_response.content = MINIMAL_YAML.encode("utf-8")
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            lot = await loader.load(yaml_inline=None, yaml_url="https://example.com/lot.yaml")
        assert lot.tender.lot_id == "6"

    @pytest.mark.asyncio
    async def test_load_url_too_large(self):
        from tilellm.modules.compliance_checker.services.yaml_requirements_loader import YamlRequirementsLoader
        loader = YamlRequirementsLoader()
        oversized = b"x" * (loader.MAX_YAML_SIZE + 1)
        mock_response = MagicMock()
        mock_response.raise_for_status = Mock()
        mock_response.content = oversized
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            with pytest.raises(ValueError, match="grande"):
                await loader.load(yaml_inline=None, yaml_url="https://example.com/big.yaml")

    @pytest.mark.asyncio
    async def test_load_malformed_yaml(self):
        from tilellm.modules.compliance_checker.services.yaml_requirements_loader import YamlRequirementsLoader
        loader = YamlRequirementsLoader()
        with pytest.raises(ValueError):
            await loader.load(yaml_inline=":::bad yaml:::", yaml_url=None)

    @pytest.mark.asyncio
    async def test_load_neither_raises(self):
        from tilellm.modules.compliance_checker.services.yaml_requirements_loader import YamlRequirementsLoader
        loader = YamlRequirementsLoader()
        with pytest.raises(ValueError, match="inline"):
            await loader.load(yaml_inline=None, yaml_url=None)


# ---------------------------------------------------------------------------
# TASK 5.3 — XlsxExtractionService.extract_cells
# ---------------------------------------------------------------------------

def _make_xlsx_bytes() -> bytes:
    """Build a minimal xlsx fixture in memory (mirrors lotti 6-9 structure)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "lotto 6"
    ws.append(["caratteristiche di conformità"])
    ws.append(["sterile"])
    ws.append(["monouso"])
    ws.append(["Criteri di valutazione", "modalità", "punt. max"])
    ws.append(["P1", "a) plasticità del tubo : max 8 punti", "variabile", "8"])
    ws.append(["P2", "b) ampiezza gamma", "proporzionale", "5"])
    ws.append(["P3", "c) certificato UNI PDR", "on/off", "1"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsxExtractionServiceCells:

    def test_extract_cells_returns_sheet_names(self):
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        svc = XlsxExtractionService(llm=None)
        sheets = svc.extract_cells(_make_xlsx_bytes())
        assert "lotto 6" in sheets

    def test_extract_cells_non_empty_rows_only(self):
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        svc = XlsxExtractionService(llm=None)
        sheets = svc.extract_cells(_make_xlsx_bytes())
        content = sheets["lotto 6"]
        assert "sterile" in content
        assert "plasticità" in content

    def test_extract_cells_multi_sheet(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "lotto 6"
        ws1.append(["sterile"])
        ws2 = wb.create_sheet("lotto 9")
        ws2.append(["monouso"])
        buf = io.BytesIO()
        wb.save(buf)
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        svc = XlsxExtractionService(llm=None)
        sheets = svc.extract_cells(buf.getvalue())
        assert "lotto 6" in sheets and "lotto 9" in sheets

    def test_extract_cells_empty_rows_excluded(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "lotto 6"
        ws.append(["sterile"])
        ws.append([None, None, None])   # empty row
        ws.append(["monouso"])
        buf = io.BytesIO()
        wb.save(buf)
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        svc = XlsxExtractionService(llm=None)
        sheets = svc.extract_cells(buf.getvalue())
        lines = [l for l in sheets["lotto 6"].splitlines() if l.strip()]
        assert len(lines) == 2


# ---------------------------------------------------------------------------
# TASK 5.4 — XlsxExtractionService.extract_requirements (LLM mocked)
# ---------------------------------------------------------------------------

def _make_llm_extraction_result(lot_id="6", lot_name="Lotto 6"):
    from tilellm.modules.compliance_checker.prompts.xlsx_extraction import (
        _LLMLotExtractionResult, _LLMTabularItem, _LLMDiscretionaryCriterion,
    )
    return _LLMLotExtractionResult(
        lot_id=lot_id,
        lot_name=lot_name,
        tabular=[
            _LLMTabularItem(id="REQ-001", text="sterile"),
            _LLMTabularItem(id="REQ-002", text="monouso"),
        ],
        discretionary=[
            _LLMDiscretionaryCriterion(id="P1", text="plasticità del tubo", mode="variabile", max_points=8),
            _LLMDiscretionaryCriterion(id="P2", text="ampiezza gamma", mode="proporzionale", max_points=5),
            _LLMDiscretionaryCriterion(id="P3", text="certificato UNI", mode="on_off", max_points=1),
        ],
        warnings=["P8: marcato human_only"],
    )


class TestXlsxExtractionServiceRequirements:

    @pytest.mark.asyncio
    async def test_extract_requirements_single_lot(self):
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        mock_structured_llm = AsyncMock()
        mock_structured_llm.ainvoke = AsyncMock(return_value=_make_llm_extraction_result())
        mock_llm = MagicMock()
        mock_llm.with_structured_output = MagicMock(return_value=mock_structured_llm)
        svc = XlsxExtractionService(llm=mock_llm)
        xlsx_bytes = _make_xlsx_bytes()
        mock_http_response = MagicMock()
        mock_http_response.raise_for_status = Mock()
        mock_http_response.content = xlsx_bytes
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_http_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            extracted = await svc.extract_requirements("https://example.com/lotti.xlsx")
        assert len(extracted) == 1
        lot = extracted[0].lot
        assert lot.tender.lot_id == "6"
        assert len(lot.requirements.tabular) == 2
        assert len(lot.requirements.discretionary) == 3
        assert extracted[0].warnings == ["P8: marcato human_only"]

    @pytest.mark.asyncio
    async def test_extract_requirements_multi_sheet(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "lotto 6"
        ws1.append(["sterile"])
        ws2 = wb.create_sheet("lotto 9")
        ws2.append(["monouso"])
        buf = io.BytesIO()
        wb.save(buf)
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        call_count = 0
        async def mock_ainvoke(_):
            nonlocal call_count
            lot_id = str(6 + call_count)
            call_count += 1
            return _make_llm_extraction_result(lot_id=lot_id, lot_name=f"Lotto {lot_id}")
        mock_structured_llm = MagicMock()
        mock_structured_llm.ainvoke = mock_ainvoke
        mock_llm = MagicMock()
        mock_llm.with_structured_output = MagicMock(return_value=mock_structured_llm)
        svc = XlsxExtractionService(llm=mock_llm)
        mock_response = MagicMock()
        mock_response.raise_for_status = Mock()
        mock_response.content = buf.getvalue()
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            lots = await svc.extract_requirements("https://example.com/multi.xlsx")
        assert len(lots) == 2

    @pytest.mark.asyncio
    async def test_extract_requirements_invalid_mode_sanitized(self):
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        from tilellm.modules.compliance_checker.prompts.xlsx_extraction import (
            _LLMLotExtractionResult, _LLMDiscretionaryCriterion,
        )
        bad_result = _LLMLotExtractionResult(
            lot_id="6", lot_name="Lotto 6",
            tabular=[],
            discretionary=[
                _LLMDiscretionaryCriterion(id="P1", text="x", mode="unknown_mode", max_points=5),
            ],
        )
        mock_structured_llm = AsyncMock()
        mock_structured_llm.ainvoke = AsyncMock(return_value=bad_result)
        mock_llm = MagicMock()
        mock_llm.with_structured_output = MagicMock(return_value=mock_structured_llm)
        svc = XlsxExtractionService(llm=mock_llm)
        mock_response = MagicMock()
        mock_response.raise_for_status = Mock()
        mock_response.content = _make_xlsx_bytes()
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            lots = await svc.extract_requirements("https://example.com/x.xlsx")
        # Invalid mode → sanitized to "variabile", criterion still included
        assert len(lots) == 1
        disc = lots[0].lot.requirements.discretionary
        assert len(disc) == 1
        assert disc[0].mode.value == "variabile"
        assert any("modalità" in w for w in lots[0].warnings)

    @pytest.mark.asyncio
    async def test_extract_requirements_zero_max_points_retained_with_warning(self):
        from tilellm.modules.compliance_checker.services.xlsx_extraction_service import XlsxExtractionService
        from tilellm.modules.compliance_checker.prompts.xlsx_extraction import (
            _LLMLotExtractionResult, _LLMDiscretionaryCriterion,
        )
        bad_result = _LLMLotExtractionResult(
            lot_id="6", lot_name="Lotto 6",
            tabular=[],
            discretionary=[
                _LLMDiscretionaryCriterion(id="P1", text="x", mode="variabile", max_points=0),
                _LLMDiscretionaryCriterion(id="P2", text="y", mode="variabile", max_points=8),
            ],
        )
        mock_structured_llm = AsyncMock()
        mock_structured_llm.ainvoke = AsyncMock(return_value=bad_result)
        mock_llm = MagicMock()
        mock_llm.with_structured_output = MagicMock(return_value=mock_structured_llm)
        svc = XlsxExtractionService(llm=mock_llm)
        mock_response = MagicMock()
        mock_response.raise_for_status = Mock()
        mock_response.content = _make_xlsx_bytes()
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get = AsyncMock(return_value=mock_response)
        with patch("httpx.AsyncClient", return_value=mock_client):
            lots = await svc.extract_requirements("https://example.com/x.xlsx")
        # P1 (max_points=0) is NOT dropped: kept with a placeholder + flagged for human review
        disc = lots[0].lot.requirements.discretionary
        assert len(disc) == 2
        p1 = next(c for c in disc if c.id == "P1")
        p2 = next(c for c in disc if c.id == "P2")
        assert p1.max_points == 1.0
        assert "PUNTEGGIO MASSIMO NON TROVATO" in p1.notes
        assert p2.max_points == 8.0
        assert any("P1" in w for w in lots[0].warnings)


# ---------------------------------------------------------------------------
# TASK 5.5 — DiscretionaryCheckService (repo + LLM mocked)
# ---------------------------------------------------------------------------

def _make_repo_mock(chunks=None, metadata=None):
    """Mock repo.get_chunks_from_repo returning a RetrievalChunksResult."""
    from tilellm.models.schemas.retrieval_schemas import RetrievalChunksResult
    chunks = chunks or ["Il prodotto è sterile e latex free."]
    metadata = metadata or [{"file_name": "offerta.pdf", "page": 5, "heading_path": "Cap.3"}]
    mock = AsyncMock()
    mock.get_chunks_from_repo = AsyncMock(
        return_value=RetrievalChunksResult(namespace="ns", chunks=chunks, metadata=metadata)
    )
    return mock


def _make_judge_response(
    coefficient=0.75,
    motivation="Buona documentazione.",
    confidence=0.85,
    source_chunk_index=1,
    evidence_text="Il prodotto è sterile",
    measured_value=None,
):
    resp = MagicMock()
    resp.content = json.dumps({
        "coefficient": coefficient,
        "measured_value": measured_value,
        "motivation": motivation,
        "confidence": confidence,
        "source_chunk_index": source_chunk_index,
        "evidence_text": evidence_text,
    })
    return resp


def _make_request_v2(min_confidence=0.6):
    return ComplianceRequestV2(
        requirements_yaml=MINIMAL_YAML,
        namespace="ns-op1",
        engine={
            "name": "pinecone", "type": "serverless",
            "apikey": "key", "vector_size": 1536, "index_name": "idx",
        },
        min_confidence=min_confidence,
    )


class TestDiscretionaryCheckService:

    def _lot_with_discretionary(self, *criteria_kwargs):
        """Build a TenderLotRequirements with given discretionary criteria."""
        from tilellm.modules.compliance_checker.models_v2 import DiscretionaryCriterion
        criteria_yaml_lines = []
        for kw in criteria_kwargs:
            lines = [
                f"    - id: {kw['id']}",
                f"      text: \"{kw['text']}\"",
                f"      mode: {kw.get('mode', 'variabile')}",
                f"      max_points: {kw.get('max_points', 8)}",
            ]
            if kw.get('human_only'):
                lines.append("      human_only: true")
            criteria_yaml_lines.extend(lines)
        criteria_block = "\n".join(criteria_yaml_lines)
        yaml_text = f"""\
tender:
  title: "Gara test"
  lot_id: "6"
  lot_name: "Lotto 6"
requirements:
  tabular: []
  discretionary:
{criteria_block}
"""
        return TenderLotRequirements.from_yaml(yaml_text)

    @pytest.mark.asyncio
    async def test_variabile_scores_correctly(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità del tubo", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=0.75))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        assert len(report.discretionary_results) == 1
        result = report.discretionary_results[0]
        assert result.coefficient == pytest.approx(0.75)
        assert result.score == pytest.approx(6.0)
        assert result.human_review_required is False

    @pytest.mark.asyncio
    async def test_variabile_coefficient_clamped(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "x", "mode": "variabile", "max_points": 10}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=1.5))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.coefficient == pytest.approx(1.0)  # clamped
        assert result.score == pytest.approx(10.0)

    @pytest.mark.asyncio
    async def test_proporzionale_extracts_value_and_flags(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P2", "text": "ampiezza gamma", "mode": "proporzionale", "max_points": 5}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(
            coefficient=None,
            measured_value="14 misure da 4.0 a 10.5 mm",
            confidence=0.85,
        ))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.coefficient is None
        assert result.score is None
        assert result.measured_value == "14 misure da 4.0 a 10.5 mm"
        assert result.human_review_required is True
        assert result.human_review_reason is not None

    @pytest.mark.asyncio
    async def test_on_off_present(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P3", "text": "certificato UNI PDR", "mode": "on_off", "max_points": 1}
        )
        repo = _make_repo_mock(
            chunks=["La società possiede la certificazione UNI PDR 125:2022."],
            metadata=[{"file_name": "cert.pdf", "page": 1}],
        )
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=1.0, confidence=0.95))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.score == pytest.approx(1.0)
        assert result.human_review_required is False

    @pytest.mark.asyncio
    async def test_on_off_absent(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P3", "text": "certificato UNI PDR", "mode": "on_off", "max_points": 1}
        )
        repo = _make_repo_mock(chunks=["Nessuna certificazione rilevante."], metadata=[{"file_name": "doc.pdf", "page": 1}])
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=0.0, confidence=0.9))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.score == pytest.approx(0.0)

    @pytest.mark.asyncio
    async def test_human_only_no_llm_call(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P8", "text": "maneggevolezza strumentario", "mode": "variabile", "max_points": 12, "human_only": True}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.human_review_required is True
        assert result.coefficient is None
        assert result.score is None
        llm.ainvoke.assert_not_called()

    @pytest.mark.asyncio
    async def test_low_confidence_triggers_human_review(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "x", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(
            coefficient=0.5, confidence=0.3  # below min_confidence=0.6
        ))
        request = _make_request_v2(min_confidence=0.6)
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.human_review_required is True
        assert "confidenza" in (result.human_review_reason or "").lower() or "confidence" in (result.human_review_reason or "").lower()

    @pytest.mark.asyncio
    async def test_no_evidence_flags_not_verifiable(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "x", "mode": "variabile", "max_points": 8}
        )
        from tilellm.models.schemas.retrieval_schemas import RetrievalChunksResult
        repo = AsyncMock()
        repo.get_chunks_from_repo = AsyncMock(
            return_value=RetrievalChunksResult(namespace="ns", chunks=[], metadata=[])
        )
        llm = AsyncMock()
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.human_review_required is True
        llm.ainvoke.assert_not_called()

    @pytest.mark.asyncio
    async def test_summary_aggregates_correctly(self):
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "x", "mode": "variabile", "max_points": 8},
            {"id": "P2", "text": "y", "mode": "proporzionale", "max_points": 5},
            {"id": "P3", "text": "z", "mode": "variabile", "max_points": 10, "human_only": True},
        )
        repo = _make_repo_mock()
        call_count = 0
        async def judge_side_effect(messages):
            nonlocal call_count
            call_count += 1
            if call_count == 1:  # P1
                return _make_judge_response(coefficient=0.75, confidence=0.9)
            else:  # P2
                return _make_judge_response(coefficient=None, measured_value="14 misure", confidence=0.8)
        llm = AsyncMock()
        llm.ainvoke = judge_side_effect
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        s = report.summary
        assert s.discretionary_total == 3
        assert s.ai_scored_count == 1          # only P1 gets a score
        assert s.ai_scored_points == pytest.approx(6.0)   # 0.75 × 8
        assert s.human_review_count == 2       # P2 (proporzionale) + P3 (human_only)

    @pytest.mark.asyncio
    async def test_reranking_disabled_uses_plain_top_k_no_rerank(self):
        """Without reranking: retrieve exactly top_k, never call the reranker."""
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock()
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=0.75))
        request = _make_request_v2()  # reranking defaults to False
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc, \
             patch("tilellm.modules.compliance_checker.services.discretionary_check_service._rerank_chunks") as mock_rerank:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            await svc.evaluate_lot(lot)
        # top_k passed to retrieval is the plain top_k (8), reranker not invoked
        qa_arg = repo.get_chunks_from_repo.call_args.args[0]
        assert qa_arg.top_k == 8
        mock_rerank.assert_not_called()

    @pytest.mark.asyncio
    async def test_reranking_enabled_oversamples_and_reranks(self):
        """With reranking: retrieve top_k×multiplier, then rerank down to top_k."""
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità", "mode": "variabile", "max_points": 8}
        )
        # repo returns 24 chunks (oversampled = top_k 8 × multiplier 3)
        repo = _make_repo_mock(
            chunks=[f"chunk {i}" for i in range(24)],
            metadata=[{"file_name": "o.pdf", "page": i} for i in range(24)],
        )
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=0.75))
        request = _make_request_v2()
        request.reranking = True
        request.reranking_multiplier = 3
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc, \
             patch("tilellm.modules.compliance_checker.services.discretionary_check_service._rerank_chunks", new_callable=AsyncMock) as mock_rerank:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            mock_rerank.return_value = (["chunk 0"], [{"file_name": "o.pdf", "page": 0}])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            await svc.evaluate_lot(lot)
        # retrieval oversamples: top_k × multiplier = 24
        qa_arg = repo.get_chunks_from_repo.call_args.args[0]
        assert qa_arg.top_k == 24
        # rerank called once, trimming back to top_k=8
        mock_rerank.assert_awaited_once()
        rerank_kwargs = mock_rerank.await_args
        assert rerank_kwargs.args[4] == 8  # final top_k

    @pytest.mark.asyncio
    async def test_reranking_failure_falls_back_gracefully(self):
        """If the reranker raises, fall back to un-reranked chunks (still judged)."""
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock(
            chunks=[f"chunk {i}" for i in range(24)],
            metadata=[{"file_name": "o.pdf", "page": i} for i in range(24)],
        )
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(coefficient=0.75, confidence=0.9))
        request = _make_request_v2()
        request.reranking = True
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc, \
             patch("tilellm.modules.compliance_checker.services.discretionary_check_service._rerank_chunks", new_callable=AsyncMock) as mock_rerank:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            mock_rerank.side_effect = RuntimeError("reranker down")
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        # judge still ran → result produced despite reranker failure
        result = report.discretionary_results[0]
        assert result.score == pytest.approx(6.0)
        llm.ainvoke.assert_awaited()

    @pytest.mark.asyncio
    async def test_citation_attributed_true_when_judge_grounds_evidence(self):
        """Judge returns a valid source_chunk_index → citation_attributed=True."""
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock(
            chunks=["Il prodotto è sterile e latex free."],
            metadata=[{"file_name": "offerta.pdf", "page": 5}],
        )
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(
            coefficient=0.75, confidence=0.9,
            source_chunk_index=1, evidence_text="Il prodotto è sterile",
        ))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        assert result.citation_attributed is True

    @pytest.mark.asyncio
    async def test_citation_not_attributed_when_judge_omits_evidence(self):
        """Judge returns empty evidence_text + source_index=0 while chunks exist → audit flag."""
        from tilellm.modules.compliance_checker.services.discretionary_check_service import DiscretionaryCheckService
        lot = self._lot_with_discretionary(
            {"id": "P1", "text": "plasticità", "mode": "variabile", "max_points": 8}
        )
        repo = _make_repo_mock(
            chunks=["Testo non correlato al criterio."],
            metadata=[{"file_name": "offerta.pdf", "page": 2}],
        )
        llm = AsyncMock()
        llm.ainvoke = AsyncMock(return_value=_make_judge_response(
            coefficient=0.5, confidence=0.9,
            source_chunk_index=0, evidence_text="",  # model failed to cite
        ))
        request = _make_request_v2()
        with patch("tilellm.modules.compliance_checker.services.discretionary_check_service.check_compliance") as mock_cc:
            from tilellm.modules.compliance_checker.models import ComplianceReport, ComplianceSummary
            mock_cc.return_value = ComplianceReport(domain="e_procurement", namespace="ns", summary=ComplianceSummary(total=0), results=[])
            svc = DiscretionaryCheckService(repo=repo, llm=llm, request=request)
            report = await svc.evaluate_lot(lot)
        result = report.discretionary_results[0]
        # score still produced, but citation flagged for the human operator
        assert result.score == pytest.approx(4.0)
        assert result.citation_attributed is False

    @pytest.mark.asyncio
    async def test_regression_v1_routes_unchanged(self):
        """V1 routes must still be importable and registered after v2 changes."""
        from tilellm.modules.compliance_checker.controllers import router
        routes = [r.path for r in router.routes]
        assert any("/check" in p for p in routes)
        assert any("/domains" in p for p in routes)


# ===========================================================================
# Aggregate report (cross-operator comparison)
# ===========================================================================

def _make_report_v2(
    namespace: str,
    lot_id: str = "6",
    lot_name: str = "Lotto 6",
    ai_points: float = 18.5,
    ai_max: float = 25.0,
    tab_total: int = 5,
    tab_compliant: int = 4,
    hr_count: int = 1,
    citation_unattr: int = 0,
) -> ComplianceReportV2:
    summary = ComplianceSummaryV2(
        tabular=TabularSummary(
            total=tab_total, compliant=tab_compliant,
            non_compliant=tab_total - tab_compliant,
        ),
        discretionary_total=3,
        ai_scored_count=2,
        ai_scored_points=ai_points,
        ai_max_scorable_points=ai_max,
        human_review_count=hr_count,
        human_review_points=5.0,
        citation_unattributed_count=citation_unattr,
    )
    disc = []
    if hr_count:
        disc.append(DiscretionaryResult(
            criterion_id="P2", criterion_text="ampiezza gamma",
            mode=DiscretionaryMode.PROPORZIONALE, max_points=5,
            motivation="m", confidence=0.8,
            human_review_required=True,
            human_review_reason="Confronto tra operatori richiesto.",
        ))
    return ComplianceReportV2(
        tender=TenderInfo(title="t", lot_id=lot_id, lot_name=lot_name),
        namespace=namespace, summary=summary,
        tabular_results=[], discretionary_results=disc,
    )


class TestAggregateReportModels:

    def test_report_source_label_optional(self):
        s = ReportSource(url="https://x.com/r.json")
        assert s.operator_label is None

    def test_request_defaults(self):
        req = AggregateReportRequest(sources=[ReportSource(url="https://x.com/r.json")])
        assert req.output_format == "md"
        assert req.synthesis_llm is False

    def test_request_requires_at_least_one_source(self):
        with pytest.raises(ValueError):
            AggregateReportRequest(sources=[])

    def test_request_format_normalized_and_validated(self):
        req = AggregateReportRequest(
            sources=[ReportSource(url="https://x.com/r.json")], output_format="DOCX",
        )
        assert req.output_format == "docx"

    def test_request_invalid_format_raises(self):
        with pytest.raises(ValueError):
            AggregateReportRequest(
                sources=[ReportSource(url="https://x.com/r.json")], output_format="xls",
            )

    def test_operator_score_carries_detail_defaults_empty(self):
        op = OperatorScore(
            operator_label="x", namespace="ns", lot_id="6", lot_name="L6",
        )
        assert op.discretionary_results == []
        assert op.tabular_results == []


class TestAggregateReportService:

    def test_derive_operator_from_namespace(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        assert AggregateReportService._derive_operator("gara-2025-lotto6-operatore-alpha") == "operatore-alpha"
        # no convention match → full namespace
        assert AggregateReportService._derive_operator("custom-ns") == "custom-ns"

    @pytest.mark.asyncio
    async def test_build_groups_by_lot_and_sorts_desc(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import patch
        reports = {
            "https://x/a.json": _make_report_v2("gara-2025-lotto6-alpha", lot_id="6", ai_points=18.5),
            "https://x/b.json": _make_report_v2("gara-2025-lotto6-beta", lot_id="6", ai_points=22.0),
            "https://x/c.json": _make_report_v2("gara-2025-lotto9-alpha", lot_id="9", ai_points=10.0),
        }
        req = AggregateReportRequest(sources=[ReportSource(url=u) for u in reports])
        svc = AggregateReportService(llm=None)
        async def fake_dl(url):
            return reports[url]
        with patch.object(svc, "_download_report", side_effect=fake_dl):
            agg = await svc.build(req)
        assert isinstance(agg, AggregateReport)
        # 2 lots
        lot_ids = {lc.lot_id for lc in agg.lots}
        assert lot_ids == {"6", "9"}
        lot6 = next(lc for lc in agg.lots if lc.lot_id == "6")
        # sorted desc by ai_scored_points → beta (22) before alpha (18.5)
        assert [o.operator_label for o in lot6.operators] == ["beta", "alpha"]

    @pytest.mark.asyncio
    async def test_explicit_label_overrides_namespace(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import patch
        report = _make_report_v2("gara-2025-lotto6-alpha", lot_id="6")
        req = AggregateReportRequest(sources=[ReportSource(url="https://x/a.json", operator_label="ACME S.p.A.")])
        svc = AggregateReportService(llm=None)
        with patch.object(svc, "_download_report", side_effect=lambda u: report):
            agg = await svc.build(req)
        assert agg.lots[0].operators[0].operator_label == "ACME S.p.A."

    @pytest.mark.asyncio
    async def test_build_populates_per_operator_detail(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import patch
        report = _make_report_v2("gara-2025-lotto6-alpha", hr_count=1)
        req = AggregateReportRequest(sources=[ReportSource(url="https://x/a.json")])
        svc = AggregateReportService(llm=None)
        with patch.object(svc, "_download_report", side_effect=lambda u: report):
            agg = await svc.build(req)
        op = agg.lots[0].operators[0]
        assert len(op.discretionary_results) == 1
        assert op.discretionary_results[0].criterion_id == "P2"

    @pytest.mark.asyncio
    async def test_human_review_items_extracted(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import patch
        report = _make_report_v2("gara-2025-lotto6-alpha", hr_count=1)
        req = AggregateReportRequest(sources=[ReportSource(url="https://x/a.json")])
        svc = AggregateReportService(llm=None)
        with patch.object(svc, "_download_report", side_effect=lambda u: report):
            agg = await svc.build(req)
        op = agg.lots[0].operators[0]
        assert op.human_review_count == 1
        assert any(it.item_id == "P2" for it in op.human_review_items)

    @pytest.mark.asyncio
    async def test_deterministic_synthesis_no_llm(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import patch
        report = _make_report_v2("gara-2025-lotto6-alpha", ai_points=18.5, ai_max=25.0)
        req = AggregateReportRequest(sources=[ReportSource(url="https://x/a.json")])  # synthesis_llm=False
        svc = AggregateReportService(llm=None)
        with patch.object(svc, "_download_report", side_effect=lambda u: report):
            agg = await svc.build(req)
        synthesis = agg.lots[0].operators[0].synthesis
        assert synthesis  # non-empty
        assert "18.5" in synthesis or "18,5" in synthesis or "74" in synthesis

    @pytest.mark.asyncio
    async def test_download_report_parses_json(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import AggregateReportService
        from unittest.mock import AsyncMock, MagicMock, Mock, patch
        report = _make_report_v2("gara-2025-lotto6-alpha")
        payload = report.model_dump(mode="json")
        resp = MagicMock()
        resp.raise_for_status = Mock()
        resp.content = json.dumps(payload).encode("utf-8")
        client = AsyncMock()
        client.__aenter__ = AsyncMock(return_value=client)
        client.__aexit__ = AsyncMock(return_value=False)
        client.get = AsyncMock(return_value=resp)
        svc = AggregateReportService(llm=None)
        with patch("httpx.AsyncClient", return_value=client):
            parsed = await svc._download_report("https://x/a.json")
        assert parsed.tender.lot_id == "6"
        assert parsed.namespace == "gara-2025-lotto6-alpha"


class TestAggregateReportRenderers:

    def _agg(self):
        from tilellm.modules.compliance_checker.models_v2 import (
            AggregateReport, LotComparison, OperatorScore, HumanReviewItem,
        )
        op = OperatorScore(
            operator_label="alpha", namespace="gara-2025-lotto6-alpha",
            lot_id="6", lot_name="Lotto 6",
            ai_scored_points=18.5, ai_max_scorable_points=25.0,
            tabular_total=5, tabular_compliant=4, tabular_compliance_rate=0.8,
            human_review_count=1, citation_unattributed_count=2,
            synthesis="Buon punteggio complessivo.",
            human_review_items=[HumanReviewItem(item_id="P2", item_type="discretionary", reason="proporzionale")],
            discretionary_results=[
                _make_discretionary_result("P1", score=6.0, coefficient=0.75),
                _make_discretionary_result(
                    "P2", mode=DiscretionaryMode.PROPORZIONALE, coefficient=None, score=None,
                    measured_value="14 misure", human_review_required=True,
                    human_review_reason="proporzionale",
                ),
            ],
        )
        return AggregateReport(
            lots=[LotComparison(lot_id="6", lot_name="Lotto 6", operators=[op])],
            total_operators=1, total_reports=1,
        )

    def test_render_markdown_contains_operator_and_scores(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import render_markdown
        md = render_markdown(self._agg())
        assert "alpha" in md
        assert "18.5" in md
        assert "Lotto 6" in md
        assert "P2" in md  # human review item surfaced

    def test_render_markdown_has_per_operator_detail_table(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import render_markdown
        md = render_markdown(self._agg())
        # the per-operator detail table lists each discretionary criterion
        assert "P1" in md
        assert "testo criterio" in md   # criterion_text from _make_discretionary_result
        assert "Modalità" in md         # detail-table header, distinct from the summary table
        # all three blocks present, in order: summary table → detail → synthesis
        assert md.index("Tabella di sintesi") < md.index("Dettaglio") < md.index("Sintesi")

    def test_detail_table_reports_evidence_document_and_page(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import render_markdown
        md = render_markdown(self._agg())
        # citation column shows where the supporting info is (doc + page), not "ok"
        assert "offerta.pdf" in md
        assert "p.5" in md
        assert "| ok |" not in md

    def test_detail_table_flags_unattributed_citation(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import render_markdown
        from tilellm.modules.compliance_checker.models_v2 import (
            AggregateReport, LotComparison, OperatorScore,
        )
        d = _make_discretionary_result("P1", score=6.0, coefficient=0.75, citation_attributed=False)
        op = OperatorScore(
            operator_label="alpha", namespace="ns", lot_id="6", lot_name="L6",
            ai_scored_points=6.0, ai_max_scorable_points=8.0,
            synthesis="s", discretionary_results=[d],
        )
        agg = AggregateReport(
            lots=[LotComparison(lot_id="6", lot_name="L6", operators=[op])],
            total_operators=1, total_reports=1,
        )
        md = render_markdown(agg)
        assert "da verificare" in md.lower()

    def test_render_report_md(self):
        from tilellm.modules.compliance_checker.services.aggregate_report_service import render_report
        content, fmt, warning = render_report(self._agg(), "md")
        assert fmt == "md"
        assert warning is None
        assert isinstance(content, (bytes, bytearray))

    def test_render_report_pdf_fallbacks_to_md_when_lib_missing(self):
        from tilellm.modules.compliance_checker.services import aggregate_report_service as mod
        from unittest.mock import patch
        # Force the pdf renderer to behave as if reportlab is unavailable
        with patch.object(mod, "render_pdf", side_effect=ImportError("no reportlab")):
            content, fmt, warning = mod.render_report(self._agg(), "pdf")
        assert fmt == "md"
        assert warning is not None and "pdf" in warning.lower()

    def test_render_report_docx_fallbacks_to_md_when_lib_missing(self):
        from tilellm.modules.compliance_checker.services import aggregate_report_service as mod
        from unittest.mock import patch
        with patch.object(mod, "render_docx", side_effect=ImportError("no python-docx")):
            content, fmt, warning = mod.render_report(self._agg(), "docx")
        assert fmt == "md"
        assert warning is not None and "docx" in warning.lower()

    def test_aggregate_route_registered(self):
        from tilellm.modules.compliance_checker.controllers import router
        routes = [r.path for r in router.routes]
        assert any("/reports/aggregate" in p for p in routes)



# ===========================================================================
# Standardized xlsx — taxonomy {Conformità/Tabellare/Discrezionale}, round-trip,
# and restituzione (output) workbook.  TDD + SOLID: taxonomy mapping is unit-tested
# in isolation; build/parse and rendering are tested per responsibility.
# ===========================================================================

import openpyxl as _openpyxl  # noqa: E402

from tilellm.modules.compliance_checker.services import xlsx_taxonomy as _tax  # noqa: E402
from tilellm.modules.compliance_checker.services.requirements_xlsx_service import (  # noqa: E402
    RequirementsXlsxService,
    select_lot,
)
from tilellm.modules.compliance_checker.services.restituzione_xlsx_service import (  # noqa: E402
    RestituzioneXlsxService,
    HEADERS as _REST_HEADERS,
    COL_OPERATOR as _REST_COL_OPERATOR,
)
from tilellm.modules.compliance_checker.services.yaml_requirements_loader import (  # noqa: E402
    YamlRequirementsLoader,
)


def _make_lot(lot_id="6", lot_name="Lotto 6") -> TenderLotRequirements:
    return TenderLotRequirements(
        tender=TenderInfo(
            title="Gara test", lot_id=lot_id, lot_name=lot_name, source_file="capitolato.xlsx"
        ),
        requirements={
            "tabular": [
                {"id": "REQ-001", "text": "sterile", "mandatory": True},
                {"id": "REQ-002", "text": "monouso opzionale", "mandatory": False},
            ],
            "discretionary": [
                {"id": "P1", "text": "plasticità del tubo", "mode": "variabile", "max_points": 8},
                {"id": "P2", "text": "ampiezza gamma", "mode": "proporzionale", "max_points": 10},
                {"id": "P3", "text": "certificato UNI", "mode": "on_off", "max_points": 1},
                {
                    "id": "P4", "text": "ergonomicità", "mode": "variabile", "max_points": 5,
                    "human_only": True, "notes": "Criterio soggettivo, non gestibile da IA",
                },
            ],
        },
    )


def _make_full_report(namespace, lot_id="6", lot_name="Lotto 6") -> ComplianceReportV2:
    tab = [
        _make_tabular_result("REQ-001", "compliant"),
        _make_tabular_result("REQ-002", "non_compliant"),
    ]
    disc = [
        _make_discretionary_result("P1", DiscretionaryMode.VARIABILE, score=6.0),
        _make_discretionary_result(
            "P2", DiscretionaryMode.PROPORZIONALE, coefficient=None, score=None,
            human_review_required=True, human_review_reason="Confronto tra operatori.",
        ),
        _make_discretionary_result(
            "P3", DiscretionaryMode.ON_OFF, max_points=1, coefficient=1.0, score=1.0,
            citation_attributed=False,
        ),
    ]
    summary = ComplianceSummaryV2.from_results(tab, disc)
    return ComplianceReportV2(
        tender=TenderInfo(title="t", lot_id=lot_id, lot_name=lot_name),
        namespace=namespace, summary=summary,
        tabular_results=tab, discretionary_results=disc,
    )


# --- Taxonomy (unit) --------------------------------------------------------

class TestXlsxTaxonomy:

    def test_normalize_type_variants(self):
        assert _tax.normalize_type("Conformità") == _tax.TYPE_CONFORMITA
        assert _tax.normalize_type("conformita") == _tax.TYPE_CONFORMITA
        assert _tax.normalize_type("TABELLARE") == _tax.TYPE_TABELLARE
        assert _tax.normalize_type(" discrezionale ") == _tax.TYPE_DISCREZIONALE
        assert _tax.normalize_type("altro") is None
        assert _tax.normalize_type(None) is None

    def test_model_type_tabular_is_conformita(self):
        assert _tax.model_type(TabularRequirementV2(id="R", text="x")) == _tax.TYPE_CONFORMITA

    def test_model_type_discretionary_modes(self):
        var = DiscretionaryCriterion(id="P", text="x", mode=DiscretionaryMode.VARIABILE, max_points=5)
        onoff = DiscretionaryCriterion(id="P", text="x", mode=DiscretionaryMode.ON_OFF, max_points=5)
        prop = DiscretionaryCriterion(id="P", text="x", mode=DiscretionaryMode.PROPORZIONALE, max_points=5)
        assert _tax.model_type(var) == _tax.TYPE_DISCREZIONALE
        assert _tax.model_type(onoff) == _tax.TYPE_TABELLARE
        assert _tax.model_type(prop) == _tax.TYPE_TABELLARE

    def test_derive_mode_from_text(self):
        assert _tax.derive_mode_from_text("...ON/OFF PUNTEGGIO 8") == DiscretionaryMode.ON_OFF
        assert _tax.derive_mode_from_text("...PROPORZIONALE 5") == DiscretionaryMode.PROPORZIONALE
        assert _tax.derive_mode_from_text("nessun indizio") is None

    def test_resolve_mode_explicit_wins(self):
        mode, warn = _tax.resolve_mode("proporzionale", "testo on/off", default=DiscretionaryMode.ON_OFF)
        assert mode == DiscretionaryMode.PROPORZIONALE and warn is None

    def test_resolve_mode_from_text_when_blank(self):
        mode, warn = _tax.resolve_mode("", "criterio ON/OFF", default=DiscretionaryMode.PROPORZIONALE)
        assert mode == DiscretionaryMode.ON_OFF and warn is None

    def test_resolve_mode_defaults_with_warning(self):
        mode, warn = _tax.resolve_mode("", "nessun indizio", default=DiscretionaryMode.ON_OFF)
        assert mode == DiscretionaryMode.ON_OFF and warn is not None


class TestExtractRequirementsOutputFormat:

    def test_default_is_yaml(self):
        assert ExtractRequirementsRequest(source="http://x/file.xlsx").output_format == "yaml"

    def test_xlsx_accepted_and_normalized(self):
        assert ExtractRequirementsRequest(source="http://x/file.xlsx", output_format="XLSX").output_format == "xlsx"

    def test_invalid_format_raises(self):
        with pytest.raises(Exception):
            ExtractRequirementsRequest(source="http://x/file.xlsx", output_format="csv")


class TestRequirementsXlsxRoundTrip:

    def test_build_produces_valid_xlsx(self):
        wb = _openpyxl.load_workbook(io.BytesIO(RequirementsXlsxService().build_workbook([_make_lot()])))
        assert len(wb.worksheets) == 1

    def test_uses_business_taxonomy_labels(self):
        content = RequirementsXlsxService().build_workbook([_make_lot()])
        wb = _openpyxl.load_workbook(io.BytesIO(content))
        flat = "\n".join(
            str(c) for row in wb.worksheets[0].iter_rows(values_only=True) for c in row if c
        )
        assert _tax.TYPE_CONFORMITA in flat
        assert _tax.TYPE_TABELLARE in flat
        assert _tax.TYPE_DISCREZIONALE in flat

    def test_roundtrip_preserves_tabular(self):
        svc = RequirementsXlsxService()
        original = _make_lot()
        parsed = svc.parse_workbook(svc.build_workbook([original]))[0]
        assert [r.model_dump() for r in parsed.requirements.tabular] == \
               [r.model_dump() for r in original.requirements.tabular]

    def test_roundtrip_preserves_discretionary(self):
        svc = RequirementsXlsxService()
        original = _make_lot()
        parsed = svc.parse_workbook(svc.build_workbook([original]))[0]
        assert [c.model_dump() for c in parsed.requirements.discretionary] == \
               [c.model_dump() for c in original.requirements.discretionary]

    def test_roundtrip_preserves_tender_metadata(self):
        svc = RequirementsXlsxService()
        parsed = svc.parse_workbook(svc.build_workbook([_make_lot()]))[0]
        assert (parsed.tender.lot_id, parsed.tender.lot_name, parsed.tender.title,
                parsed.tender.source_file) == ("6", "Lotto 6", "Gara test", "capitolato.xlsx")

    def test_roundtrip_multi_lot(self):
        svc = RequirementsXlsxService()
        lots = [_make_lot("6", "Lotto 6"), _make_lot("9", "Lotto 9")]
        parsed = svc.parse_workbook(svc.build_workbook(lots))
        assert sorted(l.tender.lot_id for l in parsed) == ["6", "9"]

    def test_parse_pure_market_sheet_derives_mode_and_id(self):
        # File following only the core market columns: no Modalità, no ID column;
        # mode must be derived from the criterion text, ids auto-generated.
        wb = _openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Lotto"); ws.cell(row=1, column=2, value="6")
        ws.cell(row=3, column=1, value="Criterio")
        ws.cell(row=3, column=2, value="Tipo criterio")
        ws.cell(row=3, column=3, value="Punteggio previsto")
        ws.cell(row=4, column=1, value="Prodotto sterile a norma")
        ws.cell(row=4, column=2, value="Conformità")
        ws.cell(row=5, column=1, value="Disponibilità versione anallergica ON/OFF PUNTEGGIO 8")
        ws.cell(row=5, column=2, value="Tabellare")
        ws.cell(row=5, column=3, value=8)
        ws.cell(row=6, column=1, value="Maggior ampiezza gamma PROPORZIONALE PUNTEGGIO 5")
        ws.cell(row=6, column=2, value="Tabellare")
        ws.cell(row=6, column=3, value=5)
        ws.cell(row=7, column=1, value="Maneggevolezza impugnatura")
        ws.cell(row=7, column=2, value="Discrezionale")
        ws.cell(row=7, column=3, value=8)
        buf = io.BytesIO(); wb.save(buf)

        lot = RequirementsXlsxService().parse_workbook(buf.getvalue())[0]
        assert len(lot.requirements.tabular) == 1
        modes = {c.id: c.mode for c in lot.requirements.discretionary}
        # auto ids: T1 (first Tabellare), T2 (second Tabellare), D1 (Discrezionale)
        assert modes["T1"] == DiscretionaryMode.ON_OFF
        assert modes["T2"] == DiscretionaryMode.PROPORZIONALE
        assert modes["D1"] == DiscretionaryMode.VARIABILE

    def test_parse_rejects_missing_punteggio_previsto(self):
        svc = RequirementsXlsxService()
        wb = _openpyxl.load_workbook(io.BytesIO(svc.build_workbook([_make_lot()])))
        ws = wb.worksheets[0]
        for row in ws.iter_rows():
            if _tax.normalize_type(row[1].value) == _tax.TYPE_TABELLARE:
                ws.cell(row=row[0].row, column=4).value = None  # clear Punteggio previsto
                break
        buf = io.BytesIO(); wb.save(buf)
        with pytest.raises(ValueError, match="Punteggio previsto"):
            svc.parse_workbook(buf.getvalue())

    def test_parse_no_header_raises(self):
        wb = _openpyxl.Workbook()
        wb.active.cell(row=1, column=1, value="qualcosa")
        buf = io.BytesIO(); wb.save(buf)
        with pytest.raises(ValueError, match="Nessun lotto valido"):
            RequirementsXlsxService().parse_workbook(buf.getvalue())

    def test_warnings_rendered_but_ignored_on_import(self):
        svc = RequirementsXlsxService()
        content = svc.build_workbook([_make_lot()], warnings_by_lot={"6": ["punteggio mancante su P9"]})
        wb = _openpyxl.load_workbook(io.BytesIO(content))
        flat = "\n".join(
            str(c) for row in wb.worksheets[0].iter_rows(values_only=True) for c in row if c
        )
        assert "punteggio mancante su P9" in flat
        assert len(svc.parse_workbook(content)[0].requirements.discretionary) == 4

    def test_human_only_and_notes_roundtrip(self):
        svc = RequirementsXlsxService()
        parsed = svc.parse_workbook(svc.build_workbook([_make_lot()]))[0]
        p4 = next(c for c in parsed.requirements.discretionary if c.id == "P4")
        assert p4.human_only is True
        assert p4.notes == "Criterio soggettivo, non gestibile da IA"


class TestSelectLot:

    def test_single_lot_no_id(self):
        assert select_lot([_make_lot("6")], None).tender.lot_id == "6"

    def test_explicit_id(self):
        assert select_lot([_make_lot("6"), _make_lot("9")], "9").tender.lot_id == "9"

    def test_multi_lot_no_id_raises(self):
        with pytest.raises(ValueError, match="più lotti"):
            select_lot([_make_lot("6"), _make_lot("9")], None)

    def test_unknown_id_raises(self):
        with pytest.raises(ValueError, match="non trovato"):
            select_lot([_make_lot("6")], "99")

    def test_empty_raises(self):
        with pytest.raises(ValueError, match="Nessun lotto"):
            select_lot([], None)


class TestComplianceRequestV2XlsxSource:

    def _base(self, **kw):
        defaults = dict(namespace="ns", engine={"name": "pinecone", "type": "serverless"})
        defaults.update(kw)
        return ComplianceRequestV2(**defaults)

    def test_xlsx_url_only_valid(self):
        assert self._base(requirements_xlsx_url="http://x/req.xlsx").requirements_xlsx_url == "http://x/req.xlsx"

    def test_xlsx_and_yaml_mutually_exclusive(self):
        with pytest.raises(Exception):
            self._base(requirements_xlsx_url="http://x/req.xlsx", requirements_yaml="tender: {}")

    def test_lot_id_field(self):
        assert self._base(requirements_xlsx_url="http://x/req.xlsx", requirements_lot_id="9").requirements_lot_id == "9"


class TestYamlRequirementsLoaderXlsx:

    @pytest.mark.asyncio
    async def test_load_from_xlsx_url(self):
        from unittest.mock import AsyncMock, MagicMock, patch
        content = RequirementsXlsxService().build_workbook([_make_lot("6"), _make_lot("9")])
        resp = MagicMock(); resp.content = content; resp.raise_for_status = MagicMock()
        client = AsyncMock()
        client.get = AsyncMock(return_value=resp)
        client.__aenter__ = AsyncMock(return_value=client)
        client.__aexit__ = AsyncMock(return_value=False)
        with patch("httpx.AsyncClient", return_value=client):
            lot = await YamlRequirementsLoader().load(
                yaml_inline=None, yaml_url=None, xlsx_url="http://x/req.xlsx", lot_id="9",
            )
        assert lot.tender.lot_id == "9"


class TestRestituzioneXlsx:

    def _sheet_rows(self, content):
        wb = _openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.worksheets[0]
        return ws, list(ws.iter_rows(values_only=True))

    def test_header_has_operatore_economico_and_market_labels(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("gara-lotto6-alpha")])
        _, rows = self._sheet_rows(content)
        header = [str(c) if c is not None else "" for c in rows[0]]
        assert header == _REST_HEADERS
        assert header[0] == _REST_COL_OPERATOR

    def test_one_row_per_criterion_single_operator(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("gara-lotto6-alpha")])
        _, rows = self._sheet_rows(content)
        data = rows[1:]
        # 2 tabular + 3 discretionary = 5 rows
        assert len([r for r in data if r[0]]) == 5

    def test_operator_label_fallbacks_to_namespace(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("gara-lotto6-alpha")])
        _, rows = self._sheet_rows(content)
        assert all(r[0] == "gara-lotto6-alpha" for r in rows[1:] if r[0])

    def test_explicit_operator_label_used(self):
        content = RestituzioneXlsxService().build_workbook(
            [_make_full_report("gara-lotto6-alpha")],
            operator_labels={"gara-lotto6-alpha": "Alpha S.p.A."},
        )
        _, rows = self._sheet_rows(content)
        assert all(r[0] == "Alpha S.p.A." for r in rows[1:] if r[0])

    def test_rows_grouped_by_criterion_then_operator(self):
        reports = [_make_full_report("ns-alpha"), _make_full_report("ns-beta")]
        content = RestituzioneXlsxService().build_workbook(reports)
        _, rows = self._sheet_rows(content)
        data = [r for r in rows[1:] if r[0]]
        # 5 criteria × 2 operators = 10 rows
        assert len(data) == 10
        # first two rows = same criterion (REQ-001), the two operators
        assert data[0][1] == data[1][1]
        assert {data[0][0], data[1][0]} == {"ns-alpha", "ns-beta"}

    def test_tabular_row_has_presenza_and_conformita(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("ns")])
        _, rows = self._sheet_rows(content)
        conf = next(r for r in rows[1:] if r[2] == _tax.TYPE_CONFORMITA)
        # Presenza filled (col 6 idx 5), Punteggio assegnato empty (col 7 idx 6)
        assert conf[5] not in (None, "")
        assert conf[6] in (None, "")

    def test_human_review_marked(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("ns")])
        _, rows = self._sheet_rows(content)
        flat = "\n".join(str(c) for r in rows for c in r if c)
        assert "REVISIONE UMANA" in flat

    def test_unattributed_citation_flagged(self):
        content = RestituzioneXlsxService().build_workbook([_make_full_report("ns")])
        _, rows = self._sheet_rows(content)
        flat = "\n".join(str(c) for r in rows for c in r if c)
        assert "da verificare" in flat

    def test_check_xlsx_route_registered(self):
        from tilellm.modules.compliance_checker.controllers import router
        routes = [r.path for r in router.routes]
        assert any(p.endswith("/check/xlsx") for p in routes)


# ===========================================================================
# YAML → xlsx export (review workbook from an existing requirements YAML)
# ===========================================================================

from tilellm.modules.compliance_checker.models_v2 import (  # noqa: E402
    RequirementsToXlsxRequest,
)
from tilellm.modules.compliance_checker.services.yaml_requirements_loader import (  # noqa: E402
    export_requirements_xlsx,
)


class TestRequirementsToXlsxRequest:

    def test_yaml_inline_only_valid(self):
        req = RequirementsToXlsxRequest(requirements_yaml=MINIMAL_YAML)
        assert req.requirements_yaml is not None

    def test_yaml_url_only_valid(self):
        req = RequirementsToXlsxRequest(requirements_yaml_url="https://x/req.yaml")
        assert req.requirements_yaml_url == "https://x/req.yaml"

    def test_both_raises(self):
        with pytest.raises(Exception):
            RequirementsToXlsxRequest(
                requirements_yaml=MINIMAL_YAML, requirements_yaml_url="https://x/req.yaml"
            )

    def test_neither_raises(self):
        with pytest.raises(Exception):
            RequirementsToXlsxRequest()


class TestExportRequirementsXlsx:

    @pytest.mark.asyncio
    async def test_export_from_yaml_inline_roundtrips(self):
        content = await export_requirements_xlsx(yaml_inline=MINIMAL_YAML, yaml_url=None)
        assert isinstance(content, (bytes, bytearray)) and len(content) > 0
        lot = RequirementsXlsxService().parse_workbook(content)[0]
        assert lot.tender.lot_id == "6"
        assert lot.requirements.tabular[0].text == "sterile"
        assert lot.requirements.discretionary[0].mode == DiscretionaryMode.VARIABILE
        assert lot.requirements.discretionary[0].max_points == 8

    @pytest.mark.asyncio
    async def test_export_uses_business_taxonomy(self):
        content = await export_requirements_xlsx(yaml_inline=MINIMAL_YAML, yaml_url=None)
        wb = _openpyxl.load_workbook(io.BytesIO(content))
        flat = "\n".join(
            str(c) for row in wb.worksheets[0].iter_rows(values_only=True) for c in row if c
        )
        assert _tax.TYPE_CONFORMITA in flat and _tax.TYPE_DISCREZIONALE in flat

    def test_to_xlsx_route_registered(self):
        from tilellm.modules.compliance_checker.controllers import router
        routes = [r.path for r in router.routes]
        assert any(p.endswith("/requirements/to-xlsx") for p in routes)


# ===========================================================================
# Bulk multi-operator evaluation (Fase 2) — proportional resolution + orchestration
# ===========================================================================

from unittest.mock import AsyncMock, patch  # noqa: E402

from tilellm.modules.compliance_checker.models_v2 import (  # noqa: E402
    BulkComplianceReport,
    BulkComplianceRequestV2,
    BulkOperatorReport,
    OperatorRef,
)
from tilellm.modules.compliance_checker.services.bulk_check_service import (  # noqa: E402
    check_compliance_v2_bulk,
    resolve_proportional,
)


def _prop_report(ns, q, cid="P2", max_points=10.0) -> ComplianceReportV2:
    d = DiscretionaryResult(
        criterion_id=cid, criterion_text="ampiezza gamma",
        mode=DiscretionaryMode.PROPORZIONALE, max_points=max_points, score=None,
        measured_value=(f"{q} misure" if q is not None else None), measured_quantity=q,
        motivation="m", confidence=0.7, human_review_required=True,
        human_review_reason="Confronto tra operatori richiesto.",
    )
    return ComplianceReportV2(
        tender=TenderInfo(title="t", lot_id="6", lot_name="Lotto 6"),
        namespace=ns, summary=ComplianceSummaryV2.from_results([], [d]),
        tabular_results=[], discretionary_results=[d],
    )


def _bulk_request(**kw):
    defaults = dict(
        requirements_yaml=MINIMAL_YAML,
        operators=[{"namespace": "ns-a", "operator_label": "Alpha"}, {"namespace": "ns-b"}],
        engine={"name": "pinecone", "type": "serverless"},
    )
    defaults.update(kw)
    return BulkComplianceRequestV2(**defaults)


class TestBulkModels:

    def test_operator_ref_label_optional(self):
        assert OperatorRef(namespace="ns").operator_label is None

    def test_requires_at_least_one_operator(self):
        with pytest.raises(Exception):
            _bulk_request(operators=[])

    def test_requirements_source_exactly_one(self):
        with pytest.raises(Exception):
            _bulk_request(requirements_yaml=MINIMAL_YAML, requirements_yaml_url="http://x/y.yaml")

    def test_requirements_source_required(self):
        with pytest.raises(Exception):
            BulkComplianceRequestV2(operators=[{"namespace": "ns"}],
                                    engine={"name": "pinecone", "type": "serverless"})

    def test_output_format_normalized(self):
        assert _bulk_request(output_format="XLSX").output_format == "xlsx"

    def test_invalid_output_format_raises(self):
        with pytest.raises(Exception):
            _bulk_request(output_format="pdf")

    def test_to_operator_request_sets_namespace_and_source(self):
        req = _bulk_request(top_k=11, min_confidence=0.7)
        op_req = req.to_operator_request("ns-x")
        assert op_req.namespace == "ns-x"
        assert op_req.requirements_yaml == MINIMAL_YAML
        assert op_req.top_k == 11
        assert op_req.min_confidence == 0.7


class TestResolveProportional:

    def test_normalizes_against_max(self):
        reports = [_prop_report("ns-a", 10.0), _prop_report("ns-b", 5.0)]
        resolve_proportional(reports)
        by_ns = {r.namespace: r.discretionary_results[0] for r in reports}
        assert by_ns["ns-a"].score == 10.0      # max → full points
        assert by_ns["ns-b"].score == 5.0       # (5/10) × 10
        assert all(d.proportional_auto for d in by_ns.values())
        assert all(d.human_review_required for d in by_ns.values())

    def test_reason_mentions_confirmation(self):
        reports = [_prop_report("ns-a", 10.0), _prop_report("ns-b", 5.0)]
        resolve_proportional(reports)
        assert "confermare" in reports[1].discretionary_results[0].human_review_reason

    def test_missing_quantity_left_unscored(self):
        reports = [_prop_report("ns-a", 10.0), _prop_report("ns-b", None)]
        resolve_proportional(reports)
        b = reports[1].discretionary_results[0]
        assert b.score is None and b.proportional_auto is False

    def test_all_missing_untouched(self):
        reports = [_prop_report("ns-a", None), _prop_report("ns-b", None)]
        resolve_proportional(reports)
        assert all(r.discretionary_results[0].score is None for r in reports)

    def test_max_operator_gets_full_points(self):
        reports = [_prop_report("ns-a", 3.0, max_points=8.0), _prop_report("ns-b", 12.0, max_points=8.0)]
        resolve_proportional(reports)
        assert reports[1].discretionary_results[0].score == 8.0


class TestBulkOrchestration:

    @pytest.mark.asyncio
    async def test_runs_each_operator_and_resolves_proportional(self):
        reports_by_ns = {"ns-a": _prop_report("ns-a", 10.0), "ns-b": _prop_report("ns-b", 5.0)}

        def fake_check(op_request):
            return reports_by_ns[op_request.namespace]

        with patch(
            "tilellm.modules.compliance_checker.services.bulk_check_service.check_compliance_v2",
            new=AsyncMock(side_effect=fake_check),
        ):
            bulk = await check_compliance_v2_bulk(_bulk_request())

        assert bulk.total_operators == 2
        assert bulk.lot_id == "6"
        labels = {o.namespace: o.operator_label for o in bulk.operators}
        assert labels == {"ns-a": "Alpha", "ns-b": "ns-b"}   # fallback to namespace
        # proportional resolved across operators
        a = next(o for o in bulk.operators if o.namespace == "ns-a").report.discretionary_results[0]
        assert a.score == 10.0 and a.proportional_auto is True

    @pytest.mark.asyncio
    async def test_bulk_xlsx_shows_da_confermare(self):
        reports_by_ns = {"ns-a": _prop_report("ns-a", 10.0), "ns-b": _prop_report("ns-b", 5.0)}

        def fake_check(op_request):
            return reports_by_ns[op_request.namespace]

        with patch(
            "tilellm.modules.compliance_checker.services.bulk_check_service.check_compliance_v2",
            new=AsyncMock(side_effect=fake_check),
        ):
            bulk = await check_compliance_v2_bulk(_bulk_request())

        labels = {o.namespace: o.operator_label for o in bulk.operators}
        content = RestituzioneXlsxService().build_workbook(
            [o.report for o in bulk.operators], operator_labels=labels
        )
        wb = _openpyxl.load_workbook(io.BytesIO(content))
        flat = "\n".join(str(c) for r in wb.worksheets[0].iter_rows(values_only=True) for c in r if c)
        assert "da confermare" in flat

    def test_bulk_route_registered(self):
        from tilellm.modules.compliance_checker.controllers import router
        routes = [r.path for r in router.routes]
        assert any(p.endswith("/check/bulk") for p in routes)
