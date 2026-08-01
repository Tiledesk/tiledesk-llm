#!/usr/bin/env python3
"""
ExcelLoader sui modelli xlsx della PA.

Due difetti visti su file di gara reali (allegato "B.2 Tracciato prodotti"):

1. **Righe di istruzioni sopra l'intestazione.** Il modello ha una o piu' righe
   di nota ("L'OE e' tenuto ad aggiungere una riga per ogni ref. offerto...")
   prima della vera riga di intestazione. `pd.read_excel` prende la riga 0 come
   header, quindi le 61 colonne reali diventano `Unnamed: 0..60`: il markdown
   prodotto ha intestazioni prive di significato e i chunk non sono
   recuperabili semanticamente. La riga di intestazione va individuata, non
   assunta — e la sua posizione cambia da file a file (riga 1 in un allegato,
   riga 3 in un altro della stessa gara).

2. **Un unico Document enorme.** Il taglio in blocchi di righe scattava solo con
   `len(df) > 100`, ma una tabella larga supera i 45.000 caratteri gia' con 29
   righe. Lasciata intera, viene poi spezzata a valle dallo splitter generico e
   **ogni chunk dopo il primo perde l'intestazione**. Tagliando nel loader,
   `to_markdown` riscrive l'intestazione in ogni blocco.
"""
import pandas as pd
import pytest
from openpyxl import Workbook

from tilellm.tools.structured_loaders import ExcelLoader


def _pa_template(path, preamble_rows, n_cols=12, n_data_rows=5, sheet="Tracciato"):
    """Foglio in stile PA: righe di nota, poi l'intestazione vera, poi i dati."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for text in preamble_rows:
        ws.append([text])
    ws.append([f"Colonna {i}" for i in range(n_cols)])
    for r in range(n_data_rows):
        ws.append([f"v{r}_{c}" for c in range(n_cols)])
    wb.save(path)
    return path


class TestHeaderRowDetection:
    def test_single_preamble_row(self, tmp_path):
        f = _pa_template(tmp_path / "b2.xlsx", ["L'OE e' tenuto ad aggiungere una riga"])
        docs = ExcelLoader(str(f)).load()
        content = "\n".join(d.page_content for d in docs)
        assert "Unnamed" not in content
        assert "Colonna 0" in content and "Colonna 11" in content

    def test_multiple_preamble_rows(self, tmp_path):
        """La posizione dell'intestazione cambia da file a file: va cercata."""
        f = _pa_template(
            tmp_path / "b2.xlsx",
            ["Si precisa che per il presente lotto...", "", "L'OE e' tenuto ad aggiungere"],
        )
        docs = ExcelLoader(str(f)).load()
        content = "\n".join(d.page_content for d in docs)
        assert "Unnamed" not in content
        assert "Colonna 0" in content

    def test_col_names_metadata_are_real_columns(self, tmp_path):
        f = _pa_template(tmp_path / "b2.xlsx", ["nota introduttiva"])
        docs = ExcelLoader(str(f)).load()
        col_names = docs[0].metadata["col_names"]
        assert "Unnamed" not in col_names
        assert "Colonna 0" in col_names

    def test_data_rows_are_preserved(self, tmp_path):
        f = _pa_template(tmp_path / "b2.xlsx", ["nota"], n_data_rows=5)
        docs = ExcelLoader(str(f)).load()
        content = "\n".join(d.page_content for d in docs)
        for r in range(5):
            assert f"v{r}_0" in content

    def test_clean_sheet_is_unaffected(self, tmp_path):
        """Nessuna riga di nota: la riga 0 e' gia' l'intestazione, niente cambia."""
        f = tmp_path / "pulito.xlsx"
        pd.DataFrame({"nome": ["Alice", "Bob"], "eta": [25, 30]}).to_excel(f, index=False)
        docs = ExcelLoader(str(f)).load()
        assert "nome" in docs[0].page_content
        assert "Alice" in docs[0].page_content
        assert "Unnamed" not in docs[0].page_content

    def test_preamble_only_sheet_does_not_crash(self, tmp_path):
        """Foglio con sole note e nessuna tabella: nessuna eccezione."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Per questo lotto non e' previsto l'allegato"])
        f = tmp_path / "vuoto.xlsx"
        wb.save(f)
        docs = ExcelLoader(str(f)).load()
        assert isinstance(docs, list)


class TestEmptyColumns:
    """Colonne interamente vuote: informazione zero, costo altissimo.

    Nei tracciati di gara reali 59 colonne su 61 risultano non compilate. Renderle
    comunque produce 58.000 caratteri di cui l'86% e' padding `nan`, che si mangia
    l'intera finestra dell'embedder (~512 token) prima di arrivare al dato utile.
    """

    def _sparse(self, tmp_path, n_cols=20, filled=("Descrizione", "Lotto")):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tracciato"
        ws.append(["L'OE e' tenuto ad aggiungere una riga"])
        cols = list(filled) + [f"Vuota_{i}" for i in range(n_cols - len(filled))]
        ws.append(cols)
        for r in range(4):
            ws.append([f"testo {r}", f"9.{r}"] + [None] * (n_cols - len(filled)))
        f = tmp_path / "sparso.xlsx"
        wb.save(f)
        return f

    def test_all_empty_columns_are_dropped(self, tmp_path):
        docs = ExcelLoader(str(self._sparse(tmp_path))).load()
        content = "\n".join(d.page_content for d in docs)
        assert "Vuota_0" not in content
        assert "Descrizione" in content and "Lotto" in content

    def test_data_survives_the_drop(self, tmp_path):
        docs = ExcelLoader(str(self._sparse(tmp_path))).load()
        content = "\n".join(d.page_content for d in docs)
        for r in range(4):
            assert f"testo {r}" in content

    def test_col_names_metadata_lists_only_kept_columns(self, tmp_path):
        docs = ExcelLoader(str(self._sparse(tmp_path))).load()
        assert "Vuota_" not in docs[0].metadata["col_names"]

    def test_partially_filled_column_is_kept(self, tmp_path):
        """Si scartano solo le colonne del tutto vuote, mai quelle con un buco."""
        wb = Workbook()
        ws = wb.active
        ws.append(["nota"])
        ws.append(["a", "b"])
        ws.append(["x", None])
        ws.append(["y", "presente"])
        f = tmp_path / "parziale.xlsx"
        wb.save(f)
        docs = ExcelLoader(str(f)).load()
        assert "presente" in docs[0].page_content


class TestWideTableChunking:
    def _wide(self, tmp_path, n_rows, n_cols):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tracciato"
        ws.append(["nota introduttiva"])
        ws.append([f"Colonna_lunga_numero_{i}" for i in range(n_cols)])
        for r in range(n_rows):
            ws.append([f"valore_abbastanza_lungo_{r}_{c}" for c in range(n_cols)])
        f = tmp_path / "largo.xlsx"
        wb.save(f)
        return f

    def test_wide_table_is_split_even_with_few_rows(self, tmp_path):
        """29 righe x 61 colonne = 45k caratteri: poche righe, documento enorme."""
        f = self._wide(tmp_path, n_rows=29, n_cols=61)
        docs = ExcelLoader(str(f)).load()
        assert len(docs) > 1, "una tabella larga va spezzata anche con poche righe"

    def test_every_chunk_repeats_the_header(self, tmp_path):
        """E' il motivo per cui il taglio va fatto qui e non a valle."""
        f = self._wide(tmp_path, n_rows=29, n_cols=61)
        docs = ExcelLoader(str(f)).load()
        for doc in docs:
            assert "Colonna_lunga_numero_0" in doc.page_content

    def test_chunks_carry_the_row_range(self, tmp_path):
        f = self._wide(tmp_path, n_rows=29, n_cols=61)
        docs = ExcelLoader(str(f)).load()
        assert all("row_range" in d.metadata for d in docs)

    def test_small_table_stays_whole(self, tmp_path):
        f = self._wide(tmp_path, n_rows=3, n_cols=4)
        docs = ExcelLoader(str(f)).load()
        assert len(docs) == 1
